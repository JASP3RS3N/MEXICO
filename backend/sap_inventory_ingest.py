"""Ingesta del inventario de SAP — SOLO LECTURA, unidireccional vía archivo plano.

Un script externo (SAP GUI Scripting / RFC, disparado por el Task Scheduler de
Windows) corre la transacción **MB52** cada hora y deja el export en una carpeta
compartida. Este módulo detecta el archivo, lo parsea y actualiza la colección
``wms_inventory_snapshots``, dejando la bitácora de cada corrida en
``wms_sap_sync_logs``.

RESTRICCIÓN DURA: aquí no existe —ni debe existir— ninguna ruta de escritura
hacia SAP. El archivo se abre en modo lectura y jamás se mueve, renombra ni
borra (la carpeta se monta como read-only en docker-compose).

MB52 sale en dos presentaciones y las dos se leen:

1. **Bloques multilínea** (la que exporta el ALV de este cliente). El
   encabezado ocupa varias líneas y cada material es un bloque separado por
   una línea en blanco::

       [MAT ]  Material Number · Material Description · Plnt · Name 1
       [CANT]  SLoc · Unrestricted · Unit · Transit · In Quality Insp. · …
       [VAL ]  Total Value · Crcy · …

   Un mismo material puede traer varios pares (CANT, VAL), uno por almacén.
   El renglón ``* Total`` del final se descarta.

2. **Fila por registro** (CSV/TSV plano o lista ALV con bordes ``|``), donde
   cada línea ya trae material, centro, almacén y cantidad.

En ambos casos el stock de libre utilización se **suma** por (parte, locación),
porque MB52 desglosa por almacén y lote.

Formatos de archivo: ``.txt`` / ``.csv`` (tabulador, ``|``, ``;`` o ``,`` — se
detecta solo) y ``.xlsx`` vía openpyxl. Los ``.xls`` viejos de SAP suelen ser
texto renombrado, así que también se intentan como texto.
"""
import csv
import glob
import io
import logging
import os
import re
import unicodedata
from typing import Optional

from config import (
    SAP_COL_DESCRIPTION,
    SAP_COL_PART_NUMBER,
    SAP_COL_PLANT,
    SAP_COL_QTY,
    SAP_COL_STORAGE_LOCATION,
    SAP_COL_UOM,
    SAP_DECIMAL_SEPARATOR,
    SAP_INVENTORY_EXPORT_PATH,
    SAP_INVENTORY_FILE_GLOB,
    SAP_INVENTORY_TENANT_SLUG,
    SAP_LOCATION_MODE,
    db,
    gen_id,
    now,
    now_iso,
    tenant_query,
)
from wms_service import minutes_between

logger = logging.getLogger("smokehouse.wms.sap")

# Alias de columnas de MB52 en inglés y español. La detección normaliza el
# encabezado (minúsculas, sin acentos ni signos) antes de comparar, así que
# "Libre utilización" y "libre utilizacion" caen en el mismo alias.
COLUMN_ALIASES = {
    "part_number": [
        "material number", "material", "matnr", "numero de parte", "num parte",
        "no parte", "part number", "part no", "articulo", "codigo material",
        "nro material", "numero material",
    ],
    "description": [
        "material description", "maktx", "descripcion", "descripcion material",
        "texto breve material", "texto breve de material", "description",
        "denominacion", "material descr",
    ],
    "plant": ["plnt", "plant", "werks", "centro", "planta"],
    "storage_location": [
        "sloc", "storage location", "lgort", "almacen", "stge loc", "stge location",
        "storage loc", "ubicacion",
    ],
    "qty": [
        "unrestricted", "labst", "libre utilizacion", "libre utiliz",
        "stock libre utilizacion", "unrestricted stock", "unrestricted use",
        "valuated unrestricted use stock", "cantidad", "stock disponible",
        "disponible", "qty", "cantidad disponible",
    ],
    "uom": [
        "unit", "base unit of measure", "meins", "bun", "umb", "unidad medida base",
        "unidad de medida", "unidad", "uom", "unit of measure", "um",
    ],
    # Solo se usa para reconocer el renglón de importes de los bloques (y para
    # afinar el separador decimal); nunca se guarda.
    "currency": ["crcy", "moneda", "waers", "curr"],
}

# Mapeo explícito por variables de entorno; si un campo viene vacío se cae a la
# autodetección por alias.
ENV_OVERRIDES = {
    "part_number": SAP_COL_PART_NUMBER,
    "description": SAP_COL_DESCRIPTION,
    "plant": SAP_COL_PLANT,
    "storage_location": SAP_COL_STORAGE_LOCATION,
    "qty": SAP_COL_QTY,
    "uom": SAP_COL_UOM,
}

REQUIRED_FIELDS = ("part_number", "plant", "qty")


# ---------------------------------------------------------------------------
# Normalización
# ---------------------------------------------------------------------------
def normalize_header(value: str) -> str:
    """Encabezado comparable: minúsculas, sin acentos, sin signos, sin dobles espacios."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().replace(".", " ").replace("_", " ").replace("-", " ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_decimal_separator(samples: list) -> str:
    """Deduce el separador decimal mirando TODA la columna de cantidad.

    Hacerlo por valor suelto es peligroso: en MB52 las cantidades llevan 3
    decimales, así que un "10,000" aislado es 10 —no diez mil— y equivocarse
    mete un error de 1000×. Mirando el archivo completo el caso se resuelve
    solo en cuanto aparece un valor con los dos separadores o con miles
    repetidos. Si el archivo es genuinamente ambiguo se asume decimal, que es
    la lectura correcta para MB52.

    Se puede forzar con SAP_DECIMAL_SEPARATOR="." o ",".
    """
    if SAP_DECIMAL_SEPARATOR in (".", ","):
        return SAP_DECIMAL_SEPARATOR

    for text in samples:
        last_dot, last_comma = text.rfind("."), text.rfind(",")
        if last_dot >= 0 and last_comma >= 0:
            # Con ambos presentes, el decimal es el que va más a la derecha.
            return "," if last_comma > last_dot else "."

    for text in samples:
        # Un separador repetido solo puede ser el de miles.
        if text.count(".") > 1:
            return ","
        if text.count(",") > 1:
            return "."

    for text in samples:
        # Un separador que no deja exactamente 3 dígitos no puede ser miles.
        for separator in (".", ","):
            position = text.rfind(separator)
            if position >= 0 and len(text) - position - 1 != 3:
                return separator

    return ","  # ambiguo: MB52 imprime 3 decimales, así que decimal


def parse_sap_number(value, decimal_separator: str = ",") -> Optional[float]:
    """Convierte una cantidad de SAP a float. None si no es un número.

    SAP imprime el signo al final ("1.234,567-") y agrupa los miles con el
    separador contrario al decimal.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = text.endswith("-")
    text = text.rstrip("-").rstrip("+").strip()
    text = text.replace(" ", "").replace("\u00a0", "")
    if not text:
        return None

    thousands = "." if decimal_separator == "," else ","
    text = text.replace(thousands, "").replace(decimal_separator, ".")

    try:
        result = float(text)
    except ValueError:
        return None
    return -result if negative else result


# ---------------------------------------------------------------------------
# Lectura de archivos
# ---------------------------------------------------------------------------
def find_latest_export(directory: str = None, pattern: str = None) -> Optional[str]:
    """Ruta del export más reciente de la carpeta, o None si no hay ninguno."""
    directory = directory or SAP_INVENTORY_EXPORT_PATH
    pattern = pattern or SAP_INVENTORY_FILE_GLOB
    if not directory or not os.path.isdir(directory):
        return None
    matches = [f for f in glob.glob(os.path.join(directory, pattern)) if os.path.isfile(f)]
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)


def _clean_cell(value) -> str:
    """Limpia una celda del ALV: quita bordes, espacios y comillas sobrantes."""
    text = str(value if value is not None else "").strip()
    return text.strip("|").strip().strip('"').strip()


def _is_decoration(cells: list) -> bool:
    """True para las líneas de adorno del ALV (---, ===, o toda la fila vacía)."""
    joined = "".join(cells).strip()
    if not joined:
        return True
    return set(joined) <= set("-=+| _")


def read_rows_from_text(content: str) -> list:
    """Filas (listas de celdas) de un export de texto.

    Detecta el delimitador entre tabulador, ``|``, punto y coma y coma contando
    cuál aparece más en las primeras líneas con contenido.

    Las líneas en blanco se conservan como ``[]``: en el formato de bloques son
    el separador entre materiales, así que tirarlas destruiría la estructura.
    Las líneas de adorno del ALV (``---``, ``===``) sí se descartan.
    """
    lines = content.splitlines()
    if not lines:
        return []

    sample = "\n".join(ln for ln in lines[:80] if ln.strip())
    counts = {d: sample.count(d) for d in ("\t", "|", ";", ",")}
    delimiter = max(counts, key=counts.get)
    if counts[delimiter] == 0:
        delimiter = "\t"

    def split(line: str) -> list:
        if delimiter == "|":
            # Lista ALV: cada línea viene rodeada y separada por barras.
            cells = [_clean_cell(c) for c in line.split("|")]
            while cells and cells[0] == "":
                cells.pop(0)
            while cells and cells[-1] == "":
                cells.pop()
            return cells
        return [_clean_cell(c) for c in next(csv.reader(io.StringIO(line), delimiter=delimiter), [])]

    rows = []
    for line in lines:
        if not line.strip():
            rows.append([])  # separador de bloque
            continue
        cells = split(line)
        if not cells or not any(cells):
            rows.append([])
        elif _is_decoration(cells):
            continue  # adorno del ALV: ni dato ni separador
        else:
            rows.append(cells)
    return rows


def read_rows_from_xlsx(path: str) -> list:
    """Filas de un .xlsx. Requiere openpyxl (está en requirements)."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = []
        for raw in sheet.iter_rows(values_only=True):
            cells = [_clean_cell(c) for c in raw]
            if not cells or not any(cells):
                rows.append([])  # separador de bloque
            elif not _is_decoration(cells):
                rows.append(cells)
        return rows
    finally:
        workbook.close()


# Codificaciones a probar, en orden. SAP corriendo sobre Windows exporta en
# cp1252 casi siempre, así que un export con "GEHÄUSE" o "CONNECTOR 90°" no es
# UTF-8 válido; leerlo como UTF-8 con errors="replace" no truena pero deja la
# descripción corrompida, y esa descripción es justo la que busca Producción.
TEXT_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")


def read_text_file(path: str) -> str:
    """Contenido del export, probando codificaciones hasta que una funcione.

    latin-1 va al final porque acepta cualquier byte: si se llega ahí, nunca
    falla, pero solo se usa cuando UTF-8 y cp1252 ya se descartaron.
    """
    for encoding in TEXT_ENCODINGS:
        try:
            with open(path, "r", encoding=encoding) as handle:
                return handle.read()
        except UnicodeDecodeError:
            continue
    # Inalcanzable en la práctica (latin-1 decodifica cualquier byte), pero
    # deja el comportamiento explícito en vez de depender de esa garantía.
    with open(path, "r", encoding="utf-8", errors="replace") as handle:
        return handle.read()


def read_rows(path: str) -> list:
    """Filas del export, eligiendo el lector por extensión.

    Un .xls de SAP casi nunca es un XLS binario real (suele ser texto o HTML
    renombrado), por eso solo .xlsx va por openpyxl.
    """
    if path.lower().endswith(".xlsx"):
        return read_rows_from_xlsx(path)
    return read_rows_from_text(read_text_file(path))


# ---------------------------------------------------------------------------
# Mapeo de columnas
# ---------------------------------------------------------------------------
def map_header_row(cells: list) -> dict:
    """Mapea UNA fila de encabezado a {campo: posición}.

    Solo devuelve los campos que esa fila resuelve: en el formato de bloques el
    encabezado está repartido en varias líneas, así que ninguna las tiene todas.
    """
    env_map = {field: normalize_header(name) for field, name in ENV_OVERRIDES.items() if name}
    normalized = [normalize_header(c) for c in cells]
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        wanted = env_map.get(field)
        if wanted:
            if wanted in normalized:
                mapping[field] = normalized.index(wanted)
            continue
        # Coincidencia exacta primero; si no, la primera que empiece igual
        # (MB52 corta encabezados largos a lo ancho de la columna).
        for alias in aliases:
            if alias in normalized:
                mapping[field] = normalized.index(alias)
                break
        else:
            for alias in aliases:
                match = next((i for i, h in enumerate(normalized) if h and h.startswith(alias)), None)
                if match is not None:
                    mapping[field] = match
                    break
    return mapping


def detect_header(rows: list) -> tuple:
    """Encuentra la fila de encabezados del formato plano (una fila por registro).

    Recorre las primeras filas porque MB52 antepone título, fecha y separadores
    antes del encabezado real. Se queda con la primera fila que resuelva todos
    los campos obligatorios; -1 si ninguna los tiene (que es justo lo que pasa
    con el formato de bloques, donde el encabezado está repartido).
    """
    for index, cells in enumerate(rows[:30]):
        mapping = map_header_row(cells)
        if all(field in mapping for field in REQUIRED_FIELDS):
            return index, mapping
    return -1, {}


# ---------------------------------------------------------------------------
# Formato de bloques multilínea (el ALV que exporta este cliente)
# ---------------------------------------------------------------------------
def detect_block_header(raw_rows: list) -> tuple:
    """Detecta el encabezado repartido en varias líneas del formato de bloques.

    Devuelve ``(primera_fila_de_datos, mapa_material, mapa_cantidad)``, o
    ``(-1, {}, {})`` si el archivo no tiene esa forma.

    La condición es que el número de parte y la cantidad vivan en líneas de
    encabezado DISTINTAS: eso es lo que distingue este formato del plano, donde
    una sola fila resuelve todo. Las posiciones salen del encabezado, no van
    fijas, así que si SAP mueve una columna el parser sigue funcionando.
    """
    material_map, quantity_map = {}, {}
    material_line = quantity_line = -1

    for index, cells in enumerate(raw_rows[:30]):
        if not any(c.strip() for c in cells):
            continue
        mapping = map_header_row(cells)
        # La línea del material trae la parte (y normalmente el centro).
        if material_line < 0 and "part_number" in mapping:
            material_map, material_line = mapping, index
            continue
        # La de cantidades trae el stock libre; el almacén y la unidad son
        # opcionales pero suelen venir aquí también.
        if quantity_line < 0 and "qty" in mapping:
            quantity_map, quantity_line = mapping, index

    if material_line < 0 or quantity_line < 0 or material_line == quantity_line:
        return -1, {}, {}
    if "plant" not in material_map and "plant" not in quantity_map:
        return -1, {}, {}

    # Los datos empiezan después de la primera línea en blanco que sigue al
    # encabezado (MB52 separa encabezado y cuerpo con un renglón vacío).
    first_data = max(material_line, quantity_line) + 1
    for index in range(first_data, min(len(raw_rows), first_data + 10)):
        if not any(c.strip() for c in raw_rows[index]):
            first_data = index + 1
            break

    return first_data, material_map, quantity_map


def split_blocks(raw_rows: list) -> list:
    """Parte las filas en bloques usando las líneas en blanco como separador."""
    blocks, current = [], []
    for cells in raw_rows:
        if any(c.strip() for c in cells):
            current.append(cells)
        elif current:
            blocks.append(current)
            current = []
    if current:
        blocks.append(current)
    return blocks


def _column(cells: list, mapping: dict, field: str) -> str:
    position = mapping.get(field)
    if position is None or position >= len(cells):
        return ""
    return cells[position].strip()


def _is_total_row(cells: list) -> bool:
    """El renglón ``* Total`` que MB52 pone al final no es un material."""
    first = next((c.strip() for c in cells if c.strip()), "")
    return first.startswith("*")


def parse_blocks(raw_rows: list, first_data: int, material_map: dict, quantity_map: dict) -> tuple:
    """Recorre los bloques y devuelve las filas de stock aplanadas.

    Dentro de un bloque, la primera línea identifica el material y las que
    siguen van en pares (cantidad, importe) — uno por almacén. Se toma la
    primera de cada par; la segunda solo aporta muestras para deducir el
    separador decimal, porque los importes siempre traen decimales.
    """
    rows_read = 0
    rows_skipped = 0
    entries = []
    value_samples = []

    for block in split_blocks(raw_rows[first_data:]):
        header = block[0]
        if _is_total_row(header):
            rows_skipped += len(block)
            continue

        part_number = _column(header, material_map, "part_number").upper()
        if not part_number:
            # Un bloque sin material (los totales del pie) no aporta stock.
            rows_skipped += len(block)
            continue

        description = _column(header, material_map, "description")
        plant = (_column(header, material_map, "plant") or _column(header, quantity_map, "plant")).upper()

        # Las líneas después del material van en pares (cantidad, importe).
        body = block[1:]
        for position in range(0, len(body), 2):
            quantity_row = body[position]
            if position + 1 < len(body):
                value_samples.append(_column(body[position + 1], quantity_map, "qty"))

            rows_read += 1
            if _is_total_row(quantity_row):
                rows_skipped += 1
                continue

            entries.append(
                {
                    "part_number": part_number,
                    "description": description,
                    "plant": plant,
                    "storage_location": _column(quantity_row, quantity_map, "storage_location"),
                    "qty_text": _column(quantity_row, quantity_map, "qty"),
                    "uom": _column(quantity_row, quantity_map, "uom"),
                }
            )

    return entries, rows_read, rows_skipped, [v for v in value_samples if v]


def build_location_code(plant: str, storage_location: str) -> str:
    """Código de locación según SAP_LOCATION_MODE: "1000" o "1000/0001"."""
    plant = (plant or "").strip().upper()
    storage_location = (storage_location or "").strip().upper()
    if SAP_LOCATION_MODE == "plant" or not storage_location:
        return plant
    return f"{plant}/{storage_location}"


def parse_export(path: str) -> tuple:
    """Parsea el export y agrega el stock por (parte, locación).

    Reconoce las dos presentaciones de MB52 —bloques multilínea y fila por
    registro— y devuelve ``(items, stats)``: items listo para guardarse como
    snapshot, stats con los conteos para el SapSyncLog.
    """
    rows = read_rows(path)

    # 1) ¿Formato de bloques? Se decide por el encabezado: si el material y la
    #    cantidad viven en líneas distintas, es bloques; si una sola fila
    #    resuelve todo, es plano.
    first_data, material_map, quantity_map = detect_block_header(rows)
    if first_data >= 0:
        layout = "bloques"
        entries, rows_read, rows_skipped, value_samples = parse_blocks(
            rows, first_data, material_map, quantity_map
        )
        columns = {**{f"material.{k}": v for k, v in sorted(material_map.items())},
                   **{f"cantidad.{k}": v for k, v in sorted(quantity_map.items())}}
    else:
        layout = "plano"
        entries, rows_read, rows_skipped, value_samples, columns = _parse_flat(rows)

    if not entries and not rows_read:
        raise ValueError(
            "No se reconocieron las columnas del export de SAP. Se necesitan al menos "
            "material, centro y stock de libre utilización. Configura SAP_COL_* en el .env "
            "con los nombres exactos de tu archivo."
        )

    # 2) El separador decimal se decide UNA vez para todo el archivo, mirando
    #    tanto las cantidades como los importes: los importes siempre traen
    #    decimales, así que resuelven el caso ambiguo (ver detect_decimal_separator).
    quantity_samples = [e["qty_text"] for e in entries if e["qty_text"]]
    decimal_separator = detect_decimal_separator(quantity_samples + value_samples)

    # 3) Agregación por (parte, locación), guardando el desglose por almacén.
    aggregated = {}
    for entry in entries:
        quantity = parse_sap_number(entry["qty_text"], decimal_separator)
        if not entry["part_number"] or not entry["plant"] or quantity is None:
            rows_skipped += 1
            continue

        storage_location = (entry["storage_location"] or "").strip().upper()
        location_code = build_location_code(entry["plant"], storage_location)
        key = (entry["part_number"], location_code)

        item = aggregated.get(key)
        if item is None:
            item = aggregated[key] = {
                "part_number": entry["part_number"],
                "description": entry["description"],
                "plant_code": entry["plant"],
                "location_code": location_code,
                "available_quantity": 0.0,
                "unit_of_measure": "",
                "storage_locations": {},
            }
        # MB52 desglosa por almacén y lote: se acumula en vez de sobrescribir.
        item["available_quantity"] += quantity
        item["description"] = item["description"] or entry["description"]
        item["unit_of_measure"] = item["unit_of_measure"] or entry["uom"].upper()
        if storage_location:
            item["storage_locations"][storage_location] = round(
                item["storage_locations"].get(storage_location, 0.0) + quantity, 3
            )

    items = []
    for item in aggregated.values():
        item["available_quantity"] = round(item["available_quantity"], 3)
        # Desglose ordenado de mayor a menor: al surtidor le sirve saber en qué
        # almacén está el grueso del material.
        item["storage_locations"] = [
            {"code": code, "quantity": quantity}
            for code, quantity in sorted(item["storage_locations"].items(), key=lambda kv: -kv[1])
        ]
        items.append(item)

    stats = {
        "rows_read": rows_read,
        "rows_skipped": rows_skipped,
        "columns": columns,
        "decimal_separator": decimal_separator,
        "layout": layout,
    }
    return items, stats


def _parse_flat(rows: list) -> tuple:
    """Lector del formato plano: una fila por registro, todas las columnas juntas."""
    data_rows = [r for r in rows if r]
    header_index, mapping = detect_header(data_rows)
    if header_index < 0:
        return [], 0, 0, [], {}

    entries = []
    for cells in data_rows[header_index + 1:]:
        if _is_total_row(cells):
            continue
        entries.append(
            {
                "part_number": _column(cells, mapping, "part_number").upper(),
                "description": _column(cells, mapping, "description"),
                "plant": _column(cells, mapping, "plant").upper(),
                "storage_location": _column(cells, mapping, "storage_location"),
                "qty_text": _column(cells, mapping, "qty"),
                "uom": _column(cells, mapping, "uom"),
            }
        )
    columns = {field: position for field, position in sorted(mapping.items())}
    return entries, len(entries), 0, [], columns


# ---------------------------------------------------------------------------
# Locaciones y persistencia
# ---------------------------------------------------------------------------
async def resolve_tenant_id() -> Optional[str]:
    """Tenant al que pertenece el export.

    Con SAP_INVENTORY_TENANT_SLUG se fija explícitamente; sin él se usa el único
    tenant existente (el caso normal de una instalación en planta). Si hay
    varios y no se configuró el slug, no se adivina.
    """
    if SAP_INVENTORY_TENANT_SLUG:
        tenant = await db.tenants.find_one({"slug": SAP_INVENTORY_TENANT_SLUG}, {"_id": 0, "id": 1})
        return tenant["id"] if tenant else None
    tenants = await db.tenants.find({}, {"_id": 0, "id": 1}).to_list(2)
    if len(tenants) == 1:
        return tenants[0]["id"]
    return None


async def ensure_location(tenant_id: str, code: str, plant_code: str) -> dict:
    """Devuelve la locación con ese código, creándola si el export trae una nueva.

    Así el supervisor no tiene que dar de alta a mano cada centro/almacén que
    aparezca en MB52; después puede renombrarla desde Ajustes.
    """
    location = await db.wms_locations.find_one(tenant_query(tenant_id, {"code": code}), {"_id": 0})
    if location:
        return location
    location = {
        "id": gen_id(),
        "tenant_id": tenant_id,
        "code": code,
        "plant_code": plant_code,
        "name": code,  # editable después desde Ajustes
        "active": True,
        "created_at": now_iso(),
        "source": "sap_import",
    }
    await db.wms_locations.insert_one(dict(location))
    logger.info("Locación creada desde el export de SAP: %s", code)
    return location


async def store_snapshots(tenant_id: str, items: list, sync_id: str, snapshot_timestamp: str) -> int:
    """Upsert del snapshot vigente por (parte, locación). Devuelve cuántos escribió.

    Se mantiene un único documento por combinación —no un histórico— para que
    la lectura del catálogo sea O(1) por locación. El histórico ligero de cada
    corrida queda en wms_sap_sync_logs.
    """
    location_cache = {}
    written = 0
    for item in items:
        code = item["location_code"]
        if code not in location_cache:
            location_cache[code] = await ensure_location(tenant_id, code, item["plant_code"])
        location = location_cache[code]

        await db.wms_inventory_snapshots.update_one(
            tenant_query(tenant_id, {"part_number": item["part_number"], "location_id": location["id"]}),
            {
                "$set": {
                    "description": item["description"],
                    "available_quantity": item["available_quantity"],
                    "unit_of_measure": item["unit_of_measure"],
                    # Desglose por almacén: el total es la suma, pero al
                    # surtidor le sirve saber en qué almacén está el material.
                    "storage_locations": item.get("storage_locations", []),
                    "plant_code": item["plant_code"],
                    "location_code": code,
                    "snapshot_timestamp": snapshot_timestamp,
                    "sync_id": sync_id,
                },
                "$setOnInsert": {
                    "id": gen_id(),
                    "tenant_id": tenant_id,
                    "part_number": item["part_number"],
                    "location_id": location["id"],
                },
            },
            upsert=True,
        )
        written += 1
    return written


async def _last_successful_log(tenant_id: str) -> Optional[dict]:
    logs = (
        await db.wms_sap_sync_logs.find(
            tenant_query(tenant_id, {"status": "success"}), {"_id": 0}
        )
        .sort("started_at", -1)
        .to_list(1)
    )
    return logs[0] if logs else None


# ---------------------------------------------------------------------------
# Corrida de sincronización
# ---------------------------------------------------------------------------
async def run_sync(tenant_id: str = None, trigger: str = "scheduler", force: bool = False) -> dict:
    """Corre una sincronización completa y devuelve el SapSyncLog resultante.

    ``force`` reprocesa el archivo aunque no haya cambiado desde la última
    corrida exitosa (lo usa el botón manual de Ajustes).
    """
    started = now()
    sync_id = gen_id()

    if tenant_id is None:
        tenant_id = await resolve_tenant_id()
    if not tenant_id:
        logger.warning("Ingesta SAP: no se pudo determinar el tenant (define SAP_INVENTORY_TENANT_SLUG).")
        return {
            "id": sync_id,
            "tenant_id": None,
            "status": "error",
            "error": "No se pudo determinar el tenant del export (define SAP_INVENTORY_TENANT_SLUG).",
            "trigger": trigger,
            "started_at": started.isoformat(),
            "finished_at": now_iso(),
            "rows_read": 0,
            "rows_upserted": 0,
            "rows_skipped": 0,
            "source_file": "",
        }

    log = {
        "id": sync_id,
        "tenant_id": tenant_id,
        "started_at": started.isoformat(),
        "finished_at": None,
        "source_file": "",
        "file_mtime": None,
        "rows_read": 0,
        "rows_upserted": 0,
        "rows_skipped": 0,
        "status": "error",
        "error": None,
        "trigger": trigger,
        "duration_ms": 0,
    }

    try:
        path = find_latest_export()
        if not path:
            raise FileNotFoundError(
                f"No se encontró ningún archivo en {SAP_INVENTORY_EXPORT_PATH} "
                f"(patrón {SAP_INVENTORY_FILE_GLOB}). Revisa que el script de SAP esté corriendo."
            )

        stat = os.stat(path)
        log["source_file"] = os.path.basename(path)
        log["file_mtime"] = stat.st_mtime

        previous = await _last_successful_log(tenant_id)
        unchanged = (
            previous
            and not force
            and previous.get("source_file") == log["source_file"]
            and previous.get("file_mtime") == stat.st_mtime
        )
        if unchanged:
            # El script de SAP no ha dejado un archivo nuevo: no reprocesamos,
            # pero sí registramos la corrida para que se vea que la app sí revisó.
            log.update(status="skipped", error=None)
        else:
            items, stats = parse_export(path)
            log["rows_read"] = stats["rows_read"]
            log["rows_skipped"] = stats["rows_skipped"]
            # Se guardan en la bitácora para que el supervisor pueda ver, sin
            # abrir el archivo, cómo lo interpretó la app.
            log["columns"] = stats["columns"]
            log["decimal_separator"] = stats["decimal_separator"]
            log["layout"] = stats["layout"]
            log["rows_upserted"] = await store_snapshots(
                tenant_id, items, sync_id, snapshot_timestamp=now_iso()
            )
            log["status"] = "success" if log["rows_upserted"] else "partial"
            if not log["rows_upserted"]:
                log["error"] = "El archivo se leyó pero no produjo ninguna fila válida."
    except Exception as exc:  # noqa: BLE001 - la ingesta nunca debe tumbar la app
        log["status"] = "error"
        log["error"] = str(exc)
        logger.warning("Ingesta SAP fallida: %s", exc)

    log["finished_at"] = now_iso()
    log["duration_ms"] = int((now() - started).total_seconds() * 1000)
    await db.wms_sap_sync_logs.insert_one(dict(log))
    if log["status"] in ("success", "partial"):
        logger.info(
            "Ingesta SAP %s: %s filas leídas, %s partes actualizadas (%s).",
            log["status"], log["rows_read"], log["rows_upserted"], log["source_file"],
        )
    return log


async def sync_all_tenants(trigger: str = "scheduler") -> list:
    """Corre la sincronización del tenant configurado. Lo usa el scheduler."""
    tenant_id = await resolve_tenant_id()
    if not tenant_id:
        logger.warning("Ingesta SAP omitida: no hay un tenant determinable.")
        return []
    return [await run_sync(tenant_id, trigger=trigger)]


# ---------------------------------------------------------------------------
# Salud de la ingesta (tarjeta del dashboard de Admin)
# ---------------------------------------------------------------------------
async def sync_health(tenant_id: str, stale_minutes: int) -> dict:
    """Estado de la ingesta: última corrida exitosa y si ya se considera caída."""
    last_success = await _last_successful_log(tenant_id)
    last_any = (
        await db.wms_sap_sync_logs.find(tenant_query(tenant_id), {"_id": 0})
        .sort("started_at", -1)
        .to_list(1)
    )
    last_run = last_any[0] if last_any else None

    minutes_since = minutes_between(last_success["started_at"]) if last_success else None
    parts_tracked = await db.wms_inventory_snapshots.count_documents(tenant_query(tenant_id))

    if last_success is None:
        status = "never"
    elif minutes_since is not None and minutes_since > stale_minutes:
        status = "stale"
    elif last_run and last_run.get("status") == "error":
        status = "degraded"
    else:
        status = "ok"

    return {
        "status": status,
        "export_path": SAP_INVENTORY_EXPORT_PATH,
        "stale_after_minutes": stale_minutes,
        "minutes_since_last_success": round(minutes_since, 1) if minutes_since is not None else None,
        "last_success": last_success,
        "last_run": last_run,
        "parts_tracked": parts_tracked,
    }
