"""WMS — dashboard de KPIs, locaciones y exportación a Excel.

Mide el desempeño del surtido: tiempos por surtidor, cumplimiento de SLA,
leaderboard, patrones del lado de Producción y surtidos parciales recurrentes
(la señal de quiebre de stock por número de parte).

Sigue el mismo formato de respuesta que los dashboards de routes_finance.py:
un dict con los agregados ya calculados, listo para pintar sin post-proceso
en el cliente.
"""
import io
import statistics
from collections import defaultdict
from datetime import timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from config import (
    WMS_CANCELLED,
    WMS_COMPLETE,
    WMS_OPEN_STATUSES,
    WMS_PARTIAL,
    WMS_PRIORITY_URGENT,
    clean,
    db,
    gen_id,
    now,
    now_iso,
    tenant_query,
)
from models_wms import LocationCreate, LocationUpdate
from security import get_tenant_id, require_owner, require_wms, user_location_id
from wms_service import decorate_request, get_wms_config, minutes_between

router = APIRouter()

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Locaciones / plantas
# ---------------------------------------------------------------------------
@router.get("/wms/locations")
async def list_locations(user: dict = Depends(require_wms)):
    """Locaciones del tenant. Un operador solo ve la suya."""
    tenant_id = get_tenant_id(user)
    extra = {}
    own = user_location_id(user)
    if own:
        extra["id"] = own
    return (
        await db.wms_locations.find(tenant_query(tenant_id, extra), {"_id": 0})
        .sort("code", 1)
        .to_list(500)
    )


@router.post("/wms/locations")
async def create_location(payload: LocationCreate, user: dict = Depends(require_owner)):
    tenant_id = get_tenant_id(user)
    if await db.wms_locations.find_one(tenant_query(tenant_id, {"code": payload.code})):
        raise HTTPException(status_code=409, detail="Ya existe una locación con ese código")
    doc = {
        "id": gen_id(),
        "tenant_id": tenant_id,
        "code": payload.code,
        "plant_code": payload.code.split("/")[0],
        "name": payload.name.strip(),
        "active": payload.active,
        "created_at": now_iso(),
        "source": "manual",
    }
    await db.wms_locations.insert_one(dict(doc))
    return clean(doc)


@router.put("/wms/locations/{location_id}")
async def update_location(
    location_id: str, payload: LocationUpdate, user: dict = Depends(require_owner)
):
    tenant_id = get_tenant_id(user)
    location = await db.wms_locations.find_one(tenant_query(tenant_id, {"id": location_id}))
    if not location:
        raise HTTPException(status_code=404, detail="Locación no encontrada")

    updates = {}
    if payload.code is not None:
        clash = await db.wms_locations.find_one(
            tenant_query(tenant_id, {"code": payload.code, "id": {"$ne": location_id}})
        )
        if clash:
            raise HTTPException(status_code=409, detail="Ya existe una locación con ese código")
        updates["code"] = payload.code
        updates["plant_code"] = payload.code.split("/")[0]
    if payload.name is not None:
        updates["name"] = payload.name.strip()
    if payload.active is not None:
        updates["active"] = payload.active

    if updates:
        await db.wms_locations.update_one(tenant_query(tenant_id, {"id": location_id}), {"$set": updates})
    return await db.wms_locations.find_one(tenant_query(tenant_id, {"id": location_id}), {"_id": 0})


# ---------------------------------------------------------------------------
# Carga de datos del periodo
# ---------------------------------------------------------------------------
def _default_range(start: str, end: str) -> tuple:
    """Rango por defecto: los últimos 30 días."""
    end = end or now_iso()
    start = start or (now() - timedelta(days=30)).isoformat()
    return start, end


def _resolve_scope(user: dict, location_id: str = None) -> dict:
    """Filtro de locación: el operador queda atado a la suya, el supervisor elige."""
    own = user_location_id(user)
    if own:
        return {"location_id": own}
    if location_id:
        return {"location_id": location_id}
    return {}


async def _requests_in_range(tenant_id: str, start: str, end: str, scope: dict) -> list:
    extra = {"requested_at": {"$gte": start, "$lte": end}, **scope}
    return await db.wms_requests.find(tenant_query(tenant_id, extra), {"_id": 0}).to_list(20000)


async def _fulfillments_in_range(tenant_id: str, start: str, end: str, scope: dict) -> list:
    extra = {"fulfilled_at": {"$gte": start, "$lte": end}, **scope}
    return await db.wms_fulfillments.find(tenant_query(tenant_id, extra), {"_id": 0}).to_list(20000)


def _avg(values: list) -> float:
    return round(sum(values) / len(values), 1) if values else 0.0


def _median(values: list) -> float:
    """Mediana: la que no se distorsiona con el turno en que todo se atoró."""
    return round(statistics.median(values), 1) if values else 0.0


def _closed_response_minutes(req: dict) -> float:
    """Minutos de respuesta de una solicitud ya cerrada."""
    return minutes_between(req.get("requested_at"), req.get("closed_at"))


# ---------------------------------------------------------------------------
# KPIs generales
# ---------------------------------------------------------------------------
@router.get("/wms/kpis")
async def wms_kpis(
    start: str = Query(None),
    end: str = Query(None),
    location_id: str = Query(None),
    user: dict = Depends(require_wms),
):
    """Resumen del periodo: volumen, tiempos, SLA y serie diaria."""
    tenant_id = get_tenant_id(user)
    start, end = _default_range(start, end)
    scope = _resolve_scope(user, location_id)
    cfg = await get_wms_config(tenant_id)

    requests = await _requests_in_range(tenant_id, start, end, scope)
    closed = [r for r in requests if r["status"] in (WMS_COMPLETE, WMS_PARTIAL)]
    open_now = [r for r in requests if r["status"] in WMS_OPEN_STATUSES]

    response_times = [_closed_response_minutes(r) for r in closed]
    within_sla = [m for m in response_times if m <= cfg["sla_minutes"]]

    requested_qty = sum(float(r.get("quantity_requested", 0) or 0) for r in closed)
    fulfilled_qty = sum(float(r.get("quantity_fulfilled_total", 0) or 0) for r in closed)

    # Serie diaria de volumen y tiempo promedio, para la gráfica del dashboard.
    by_day = defaultdict(lambda: {"requests": 0, "closed": 0, "minutes": []})
    for r in requests:
        day = (r.get("requested_at") or "")[:10]
        by_day[day]["requests"] += 1
        if r["status"] in (WMS_COMPLETE, WMS_PARTIAL):
            by_day[day]["closed"] += 1
            by_day[day]["minutes"].append(_closed_response_minutes(r))
    series = [
        {
            "date": day,
            "requests": values["requests"],
            "closed": values["closed"],
            "avg_minutes": _avg(values["minutes"]),
        }
        for day, values in sorted(by_day.items())
        if day
    ]

    return {
        "range": {"start": start, "end": end},
        "config": cfg,
        "totals": {
            "requests": len(requests),
            "closed": len(closed),
            "open": len(open_now),
            "cancelled": sum(1 for r in requests if r["status"] == WMS_CANCELLED),
            "urgent": sum(1 for r in requests if r.get("priority") == WMS_PRIORITY_URGENT),
            "partial": sum(1 for r in requests if r["status"] == WMS_PARTIAL),
        },
        "response_minutes": {
            "avg": _avg(response_times),
            "median": _median(response_times),
            "max": round(max(response_times), 1) if response_times else 0.0,
        },
        "sla": {
            "target_minutes": cfg["sla_minutes"],
            "within": len(within_sla),
            "outside": len(response_times) - len(within_sla),
            "pct_within": round(100 * len(within_sla) / len(response_times), 1) if response_times else 0.0,
        },
        "quantities": {
            "requested": round(requested_qty, 3),
            "fulfilled": round(fulfilled_qty, 3),
            "fill_rate_pct": round(100 * fulfilled_qty / requested_qty, 1) if requested_qty else 0.0,
        },
        "series": series,
    }


# ---------------------------------------------------------------------------
# Leaderboard de surtidores
# ---------------------------------------------------------------------------
@router.get("/wms/leaderboard")
async def wms_leaderboard(
    start: str = Query(None),
    end: str = Query(None),
    location_id: str = Query(None),
    user: dict = Depends(require_wms),
):
    """Ranking de almacén: quién surte más rápido y quién surte más volumen.

    Se reportan dos tiempos por persona: el de respuesta total (desde que
    Producción solicitó, que incluye la espera en cola) y el propio de surtido
    (desde que tomó la solicitud), que es el que realmente depende del surtidor.
    """
    tenant_id = get_tenant_id(user)
    start, end = _default_range(start, end)
    scope = _resolve_scope(user, location_id)
    cfg = await get_wms_config(tenant_id)

    fulfillments = await _fulfillments_in_range(tenant_id, start, end, scope)

    by_person = defaultdict(
        lambda: {
            "fulfillments": 0,
            "quantity": 0.0,
            "response_minutes": [],
            "handling_minutes": [],
            "requests": set(),
        }
    )
    for f in fulfillments:
        person = by_person[f.get("fulfilled_by_user_id") or "desconocido"]
        person["name"] = f.get("fulfilled_by_name") or "Desconocido"
        person["fulfillments"] += 1
        person["quantity"] += float(f.get("quantity_fulfilled", 0) or 0)
        person["response_minutes"].append(float(f.get("time_to_fulfill_minutes", 0) or 0))
        person["handling_minutes"].append(float(f.get("time_since_claim_minutes", 0) or 0))
        person["requests"].add(f.get("request_id"))

    rows = []
    for user_id, data in by_person.items():
        response = data["response_minutes"]
        within = [m for m in response if m <= cfg["sla_minutes"]]
        rows.append(
            {
                "user_id": user_id,
                "name": data["name"],
                "fulfillments": data["fulfillments"],
                "requests_served": len(data["requests"]),
                "quantity_fulfilled": round(data["quantity"], 3),
                "avg_response_minutes": _avg(response),
                "median_response_minutes": _median(response),
                "avg_handling_minutes": _avg(data["handling_minutes"]),
                "median_handling_minutes": _median(data["handling_minutes"]),
                "pct_within_sla": round(100 * len(within) / len(response), 1) if response else 0.0,
            }
        )

    # Orden por volumen: es el criterio que no castiga a quien tomó las difíciles.
    rows.sort(key=lambda r: (-r["requests_served"], r["median_response_minutes"]))
    return {"range": {"start": start, "end": end}, "sla_minutes": cfg["sla_minutes"], "rows": rows}


# ---------------------------------------------------------------------------
# Lado de Producción: quién solicita y con qué patrón
# ---------------------------------------------------------------------------
@router.get("/wms/requesters")
async def wms_requesters(
    start: str = Query(None),
    end: str = Query(None),
    location_id: str = Query(None),
    user: dict = Depends(require_wms),
):
    """Actividad por solicitante, con el peso de las urgentes.

    Una proporción alta de urgentes repetidas en la misma persona/parte suele
    apuntar a planeación de turno, no a lentitud de almacén — por eso se
    reporta aparte del leaderboard.
    """
    tenant_id = get_tenant_id(user)
    start, end = _default_range(start, end)
    scope = _resolve_scope(user, location_id)

    requests = await _requests_in_range(tenant_id, start, end, scope)

    by_person = defaultdict(
        lambda: {"requests": 0, "urgent": 0, "parts": defaultdict(int), "days": set(), "minutes": []}
    )
    for r in requests:
        person = by_person[r.get("requested_by_user_id") or "desconocido"]
        person["name"] = r.get("requested_by_name") or "Desconocido"
        person["requests"] += 1
        if r.get("priority") == WMS_PRIORITY_URGENT:
            person["urgent"] += 1
            person["parts"][r["part_number"]] += 1
        person["days"].add((r.get("requested_at") or "")[:10])
        if r["status"] in (WMS_COMPLETE, WMS_PARTIAL):
            person["minutes"].append(_closed_response_minutes(r))

    rows = []
    for user_id, data in by_person.items():
        # Urgentes repetidas sobre la MISMA parte: el patrón que vale la pena mirar.
        repeated = sorted(
            ({"part_number": part, "urgent_count": count} for part, count in data["parts"].items() if count >= 2),
            key=lambda item: -item["urgent_count"],
        )
        rows.append(
            {
                "user_id": user_id,
                "name": data["name"],
                "requests": data["requests"],
                "urgent": data["urgent"],
                "pct_urgent": round(100 * data["urgent"] / data["requests"], 1) if data["requests"] else 0.0,
                "active_days": len([d for d in data["days"] if d]),
                "requests_per_active_day": round(
                    data["requests"] / max(len([d for d in data["days"] if d]), 1), 1
                ),
                "avg_wait_minutes": _avg(data["minutes"]),
                "repeated_urgent_parts": repeated[:5],
            }
        )

    rows.sort(key=lambda r: -r["requests"])
    return {"range": {"start": start, "end": end}, "rows": rows}


# ---------------------------------------------------------------------------
# Surtidos parciales por número de parte (señal de quiebre de stock)
# ---------------------------------------------------------------------------
@router.get("/wms/partial-fulfillments")
async def wms_partial_fulfillments(
    start: str = Query(None),
    end: str = Query(None),
    location_id: str = Query(None),
    min_occurrences: int = Query(1, ge=1),
    user: dict = Depends(require_wms),
):
    """Comparativo solicitado vs surtido por parte.

    Una parte que se surte parcial una y otra vez es un quiebre de stock
    recurrente, no un problema de la persona que la surtió.
    """
    tenant_id = get_tenant_id(user)
    start, end = _default_range(start, end)
    scope = _resolve_scope(user, location_id)

    requests = await _requests_in_range(tenant_id, start, end, scope)
    closed = [r for r in requests if r["status"] in (WMS_COMPLETE, WMS_PARTIAL)]

    by_part = defaultdict(
        lambda: {"requests": 0, "partials": 0, "requested": 0.0, "fulfilled": 0.0, "description": "", "unit": ""}
    )
    for r in closed:
        part = by_part[r["part_number"]]
        part["description"] = part["description"] or r.get("description", "")
        part["unit"] = part["unit"] or r.get("unit_of_measure", "")
        part["requests"] += 1
        part["requested"] += float(r.get("quantity_requested", 0) or 0)
        part["fulfilled"] += float(r.get("quantity_fulfilled_total", 0) or 0)
        if r["status"] == WMS_PARTIAL:
            part["partials"] += 1

    rows = [
        {
            "part_number": part_number,
            "description": data["description"],
            "unit_of_measure": data["unit"],
            "requests": data["requests"],
            "partial_count": data["partials"],
            "pct_partial": round(100 * data["partials"] / data["requests"], 1) if data["requests"] else 0.0,
            "quantity_requested": round(data["requested"], 3),
            "quantity_fulfilled": round(data["fulfilled"], 3),
            "shortfall": round(max(data["requested"] - data["fulfilled"], 0), 3),
            "fill_rate_pct": round(100 * data["fulfilled"] / data["requested"], 1) if data["requested"] else 0.0,
        }
        for part_number, data in by_part.items()
        if data["partials"] >= min_occurrences
    ]
    rows.sort(key=lambda r: (-r["partial_count"], -r["shortfall"]))
    return {"range": {"start": start, "end": end}, "rows": rows}


# ---------------------------------------------------------------------------
# Exportación a Excel
# ---------------------------------------------------------------------------
DETAIL_HEADERS = [
    "Folio", "Número de parte", "Descripción", "Cantidad solicitada", "Unidad",
    "Cantidad surtida", "Pendiente", "Solicitado por", "Surtido por",
    "Fecha/hora solicitud", "Fecha/hora surtido", "Minutos de respuesta",
    "Dentro de SLA", "Prioridad", "Status", "Locación", "Planta", "Notas",
]

KPI_HEADERS = [
    "Persona", "Rol", "Solicitudes", "Surtidos registrados", "Cantidad surtida",
    "Minutos promedio", "Minutos mediana", "% dentro de SLA", "Urgentes",
]

STATUS_LABELS = {
    "pendiente": "Pendiente",
    "en_proceso": "En proceso",
    "surtido_parcial": "Surtido parcial",
    "surtido_completo": "Surtido completo",
    "cancelado": "Cancelado",
}


def _local_datetime(iso_value: str) -> str:
    """ISO → 'YYYY-MM-DD HH:MM' legible en Excel. Vacío si no hay dato."""
    if not iso_value:
        return ""
    return str(iso_value)[:16].replace("T", " ")


@router.get("/wms/export/excel")
async def export_excel(
    start: str = Query(None, alias="from", description="YYYY-MM-DD"),
    end: str = Query(None, alias="to", description="YYYY-MM-DD"),
    location_id: str = Query(None),
    user: dict = Depends(require_owner),
):
    """Descarga el detalle del periodo + una hoja de KPIs por persona.

    Genera el .xlsx en memoria con openpyxl (no se toca el disco ni se deja
    ningún archivo temporal en el contenedor).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    tenant_id = get_tenant_id(user)
    # Los parámetros vienen como fecha; se expanden al día completo.
    start_iso = f"{start}T00:00:00+00:00" if start else None
    end_iso = f"{end}T23:59:59+00:00" if end else None
    start_iso, end_iso = _default_range(start_iso, end_iso)
    scope = _resolve_scope(user, location_id)
    cfg = await get_wms_config(tenant_id)

    requests = await _requests_in_range(tenant_id, start_iso, end_iso, scope)
    requests.sort(key=lambda r: r.get("requested_at") or "")

    # Último surtidor por solicitud, para la columna "Surtido por".
    fulfillments = await db.wms_fulfillments.find(
        tenant_query(tenant_id, {"request_id": {"$in": [r["id"] for r in requests]}}), {"_id": 0}
    ).to_list(20000)
    last_fulfiller = {}
    for f in sorted(fulfillments, key=lambda item: item.get("fulfilled_at") or ""):
        last_fulfiller[f["request_id"]] = f

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="B45309")  # ámbar, el acento de la app

    def write_header(sheet, headers):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"

    def autosize(sheet, headers, max_width=48):
        for index, header in enumerate(headers, start=1):
            longest = len(str(header))
            for cell in sheet[get_column_letter(index)][1:]:
                longest = max(longest, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(index)].width = min(longest + 2, max_width)

    # --- Hoja 1: detalle de solicitudes ---
    detail = workbook.active
    detail.title = "Solicitudes"
    write_header(detail, DETAIL_HEADERS)

    for req in requests:
        decorated = decorate_request(req, cfg)
        fulfillment = last_fulfiller.get(req["id"])
        is_closed = not decorated["is_open"]
        detail.append(
            [
                req.get("folio", ""),
                req.get("part_number", ""),
                req.get("description", ""),
                float(req.get("quantity_requested", 0) or 0),
                req.get("unit_of_measure", ""),
                float(req.get("quantity_fulfilled_total", 0) or 0),
                decorated["quantity_pending"],
                req.get("requested_by_name", ""),
                (fulfillment or {}).get("fulfilled_by_name", "") or req.get("claimed_by_name", "") or "",
                _local_datetime(req.get("requested_at")),
                _local_datetime(req.get("closed_at")),
                decorated["minutes_elapsed"] if is_closed else "",
                ("Sí" if decorated["within_sla"] else "No") if is_closed else "",
                req.get("priority", ""),
                STATUS_LABELS.get(req.get("status"), req.get("status", "")),
                req.get("location_name") or req.get("plant_code", ""),
                req.get("plant_code", ""),
                req.get("notes", ""),
            ]
        )
    autosize(detail, DETAIL_HEADERS)

    # --- Hoja 2: KPIs por persona ---
    kpis = workbook.create_sheet("KPIs por persona")
    write_header(kpis, KPI_HEADERS)

    warehouse_rows = (
        await wms_leaderboard(start=start_iso, end=end_iso, location_id=location_id, user=user)
    )["rows"]
    for row in warehouse_rows:
        kpis.append(
            [
                row["name"], "Almacén", "", row["fulfillments"], row["quantity_fulfilled"],
                row["avg_response_minutes"], row["median_response_minutes"], row["pct_within_sla"], "",
            ]
        )

    production_rows = (
        await wms_requesters(start=start_iso, end=end_iso, location_id=location_id, user=user)
    )["rows"]
    for row in production_rows:
        kpis.append(
            [
                row["name"], "Producción", row["requests"], "", "",
                row["avg_wait_minutes"], "", "", row["urgent"],
            ]
        )
    autosize(kpis, KPI_HEADERS)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"wms_{start_iso[:10]}_{end_iso[:10]}.xlsx"
    return StreamingResponse(
        stream,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
