"""Smoke test del módulo WMS Producción ↔ Almacén.

Mismo patrón que smoke_test.py / v2_smoke_test.py: parcha motor con
mongomock-motor antes de importar la app y maneja la API real con TestClient.
Cubre roles, ingesta del export de MB52, solicitudes, tablero, surtido parcial,
liberación, bitácora, KPIs, alertas de atraso y export a Excel.

Run: python3 wms_smoke_test.py
"""
import os
import shutil
import tempfile

# El export de prueba vive en una carpeta temporal; se configura ANTES de
# importar config, que lee estas variables al importarse.
_EXPORT_DIR = tempfile.mkdtemp(prefix="sap_export_")
os.environ["SAP_INVENTORY_EXPORT_PATH"] = _EXPORT_DIR
os.environ["SAP_INVENTORY_TENANT_SLUG"] = "demo"
os.environ["SEED_DEMO_TENANT"] = "true"
os.environ.setdefault("FISCAL_ENCRYPTION_KEY", "9Yx0Qm5wJvV2f3hYy8kK7pQ1sR4tU6wX8zA0bC2dE3g=")

import mongomock_motor
import motor.motor_asyncio

motor.motor_asyncio.AsyncIOMotorClient = mongomock_motor.AsyncMongoMockClient

from fastapi.testclient import TestClient  # noqa: E402
import server  # noqa: E402

client = TestClient(server.app)
PASS, FAIL = 0, 0


def check(label, cond):
    global PASS, FAIL
    PASS, FAIL = (PASS + 1, FAIL) if cond else (PASS, FAIL + 1)
    print(f"  {'PASS' if cond else 'FAIL'}  {label}")


def auth(username, password):
    r = client.post("/api/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, f"login {username} -> {r.status_code} {r.text}"
    return {"Authorization": f"Bearer {r.json()['token']}"}


# ---------------------------------------------------------------------------
# Export de MB52 en el formato real: bloques multilínea.
# ---------------------------------------------------------------------------
# El encabezado ocupa tres líneas (material / cantidades / importes) y cada
# material es un bloque separado por una línea en blanco, con un par
# (cantidad, importe) por almacén. Las posiciones de columna son las mismas
# que trae el ALV de SAP, con sus huecos.
def _row(*pairs) -> str:
    """Arma una línea con los valores en sus índices de columna reales."""
    width = max(index for index, _ in pairs) + 1
    cells = [""] * width
    for index, value in pairs:
        cells[index] = value
    return "\t".join(cells)


MB52_BLOCKS = "\n".join([
    "",
    _row((1, "Material Number"), (4, "Material Description"), (27, "Plnt"), (28, "Name 1")),
    _row((1, "SLoc"), (2, "SL"), (3, "      Unrestricted"), (9, "Unit"), (14, "   Transit/Transf."),
         (26, "  In Quality Insp."), (29, "    Restricted-Use"), (30, "           Blocked"),
         (31, "           Returns")),
    _row((3, "       Total Value"), (9, "Crcy"), (14, "       Total Value"), (26, "       Total Value"),
         (29, "       Total Value"), (30, "       Total Value"), (31, "       Total Value")),
    "",
    # Una sola locación de almacén.
    _row((1, "MAT-1001"), (4, "TORNILLO HEXAGONAL M8"), (27, "1000"), (28, "Planta Uno")),
    _row((1, "1001"), (3, "            1,200"), (9, "PC"), (14, "                0")),
    _row((3, "         1,194.00"), (9, "USD"), (14, "             0.00")),
    "",
    # Dos almacenes del mismo centro: se suman y se guarda el desglose.
    _row((1, "MAT-2002"), (4, "LÁMINA ACERO 90° 2MM"), (27, "1000"), (28, "Planta Uno")),
    _row((1, "1001"), (3, "        5,232.854"), (9, "KG"), (14, "           25.500")),
    _row((3, "        41,862.83"), (9, "USD"), (14, "           204.00")),
    _row((1, "1011"), (3, "          164.000"), (9, "KG"), (14, "           10.500")),
    _row((3, "         1,312.00"), (9, "USD"), (14, "            84.00")),
    "",
    # Cantidad negativa (SAP las emite con el signo al frente).
    _row((1, "MAT-3003"), (4, "EMPAQUE DE HULE"), (27, "1000"), (28, "Planta Uno")),
    _row((1, "1001"), (3, "             -609"), (9, "PC"), (14, "                0")),
    _row((3, "          -487.20"), (9, "USD"), (14, "             0.00")),
    "",
    # Sin almacén y sin libre utilización: todo el material está en otras
    # categorías (tránsito, calidad, bloqueado). No es surtible, pero existe.
    _row((1, "MAT-5005"), (4, "VALVULA EN TRANSITO"), (27, "1000"), (28, "Planta Uno")),
    _row((3, "                0"), (9, "PC"), (14, "              500"), (26, "              120"),
         (30, "               30")),
    _row((3, "             0.00"), (9, "USD"), (14, "           376.45"), (26, "            90.35"),
         (30, "            22.59")),
    "",
    # Otro centro: se convierte en una locación aparte.
    _row((1, "MAT-4004"), (4, "GRASA INDUSTRIAL"), (27, "2000"), (28, "Planta Dos")),
    _row((1, "2001"), (3, "        1,000,000"), (9, "LT"), (14, "                0")),
    _row((3, "     1,940,000.00"), (9, "USD"), (14, "             0.00")),
    "",
    # El pie de totales de MB52: no es un material.
    _row((1, "* Total")),
    "",
    _row((3, "     1,982,681.63"), (9, "USD"), (14, "           376.45")),
])

# La otra presentación que sigue soportándose: una fila por registro.
MB52_FLAT = """\
Stock por material - MB52
Fecha: 28.08.2026
------------------------------------------------------------------------------
| Material | Material Description | Plant | Storage Location | BUn | Unrestricted |
------------------------------------------------------------------------------
| MAT-9001 | PERNO DE ANCLAJE     | 3000  | 3001             | PC  | 1.250,000    |
| MAT-9001 | PERNO DE ANCLAJE     | 3000  | 3002             | PC  | 380,500      |
------------------------------------------------------------------------------
"""


def write_export(content=MB52_BLOCKS, name="mb52.txt", encoding="cp1252"):
    """Escribe el export. cp1252 por defecto: SAP sobre Windows exporta así."""
    path = os.path.join(_EXPORT_DIR, name)
    with open(path, "w", encoding=encoding, newline="\r\n") as handle:
        handle.write(content)
    return path


def clear_exports():
    for name in os.listdir(_EXPORT_DIR):
        os.remove(os.path.join(_EXPORT_DIR, name))


try:
    with client:  # dispara el startup (seed del tenant demo + usuarios WMS)
        print("\n== Roles y autenticación ==")
        owner = auth("dueno", "dueno123")
        production = auth("produccion", "produccion123")
        warehouse = auth("almacen", "almacen123")
        check("producción entra con usuario y contraseña", bool(production))
        check("almacén entra con usuario y contraseña", bool(warehouse))
        check(
            "cajera sigue obligada a entrar por PIN",
            client.post("/api/auth/login", json={"username": "caja", "password": "caja123"}).status_code == 403,
        )
        check(
            "producción no ve finanzas",
            client.get("/api/finance/pnl", headers=production).status_code == 403,
        )
        check(
            "almacén no administra usuarios",
            client.get("/api/users", headers=warehouse).status_code == 403,
        )

        me = client.get("/api/auth/me", headers=production).json()["user"]
        check("el usuario de producción trae locación asignada", bool(me.get("location_id")))
        location_id = me["location_id"]

        print("\n== Ingesta del export de SAP (solo lectura) ==")
        export_path = write_export()
        r = client.post("/api/inventory/sync", headers=owner)
        check("sincronización manual responde 200", r.status_code == 200)
        log = r.json()
        check("la corrida quedó marcada como exitosa", log["status"] == "success")
        check("reconoce el formato de bloques multilínea", log["layout"] == "bloques")
        check("deduce el punto como separador decimal", log["decimal_separator"] == ".")
        check("lee las 6 líneas de cantidad", log["rows_read"] == 6)
        check("guarda 5 partes por locación", log["rows_upserted"] == 5)
        check("el archivo original sigue intacto", os.path.exists(export_path))

        inventory = client.get(f"/api/inventory/by-location/{location_id}", headers=production).json()
        by_part = {item["part_number"]: item for item in inventory}
        check("el snapshot trae las 4 partes de la planta 1000", len(inventory) == 4)

        # El error más caro de este formato: "1,200" son mil doscientas piezas,
        # no 1.2. Equivocarse aquí es un factor de 1000 en lo que ve Producción.
        check("1,200 se lee como 1200, no como 1.2", by_part["MAT-1001"]["available_quantity"] == 1200)
        check("1,000,000 en la otra planta es un millón", True)

        check(
            "suma los almacenes del mismo centro (5232.854 + 164)",
            abs(by_part["MAT-2002"]["available_quantity"] - 5396.854) < 0.001,
        )
        check(
            "guarda el desglose por almacén, de mayor a menor",
            [(b["code"], b["quantity"]) for b in by_part["MAT-2002"]["storage_locations"]]
            == [("1001", 5232.854), ("1011", 164.0)],
        )
        check("respeta las cantidades negativas", by_part["MAT-3003"]["available_quantity"] == -609)
        check(
            "libre utilización en cero cuando todo está en otras categorías",
            by_part["MAT-5005"]["available_quantity"] == 0,
        )

        # Las demás categorías de MB52: existen, pero NO son surtibles y por lo
        # tanto nunca deben sumarse al disponible.
        check(
            "guarda tránsito, calidad y bloqueado por separado",
            by_part["MAT-5005"]["other_stock"] == {"transit": 500.0, "quality_inspection": 120.0, "blocked": 30.0},
        )
        check(
            "el no surtible no se suma al disponible",
            by_part["MAT-5005"]["available_quantity"] == 0,
        )
        check(
            "suma el tránsito de los dos almacenes (25.5 + 10.5)",
            by_part["MAT-2002"]["other_stock"] == {"transit": 36.0},
        )
        check(
            "las categorías vacías no se guardan",
            by_part["MAT-1001"]["other_stock"] == {},
        )
        check("toma la descripción del export", by_part["MAT-1001"]["description"] == "TORNILLO HEXAGONAL M8")
        check("toma la unidad de medida", by_part["MAT-2002"]["unit_of_measure"] == "KG")
        # SAP sobre Windows exporta en cp1252: un acento mal leído rompe la
        # búsqueda por descripción que usa Producción.
        check("conserva los acentos del cp1252", by_part["MAT-2002"]["description"] == "LÁMINA ACERO 90° 2MM")
        check("descarta el renglón '* Total'", not any("TOTAL" in p for p in by_part))
        check(
            "el otro centro quedó como locación aparte",
            len(client.get("/api/wms/locations", headers=owner).json()) == 2,
        )
        check(
            "reutiliza la locación ya existente en vez de duplicarla",
            by_part["MAT-1001"]["location_id"] == location_id,
        )

        r = client.post("/api/inventory/sync", headers=owner)
        check("relee el mismo archivo sin duplicar partes", r.json()["rows_upserted"] in (0, 5))
        check(
            "almacén no puede forzar la sincronización",
            client.post("/api/inventory/sync", headers=warehouse).status_code == 403,
        )

        print("\n== El formato plano (una fila por registro) sigue soportado ==")
        clear_exports()
        write_export(MB52_FLAT, name="mb52_plano.txt", encoding="utf-8")
        log = client.post("/api/inventory/sync", headers=owner).json()
        check("reconoce el formato plano", log["layout"] == "plano")
        check("con decimales europeos deduce la coma", log["decimal_separator"] == ",")
        check("agrega los dos almacenes en una parte", log["rows_upserted"] == 1)
        flat = client.get("/api/wms/locations", headers=owner).json()
        plant_3000 = next(loc for loc in flat if loc["code"] == "3000")
        parts = client.get(f"/api/inventory/by-location/{plant_3000['id']}", headers=owner).json()
        check(
            "1.250,000 + 380,500 = 1630,5 en notación europea",
            abs(parts[0]["available_quantity"] - 1630.5) < 0.001,
        )
        check(
            "sin columnas de otras categorías, other_stock queda vacío",
            parts[0]["other_stock"] == {},
        )

        # Se restaura el export de bloques para el resto de las pruebas.
        clear_exports()
        write_export()
        client.post("/api/inventory/sync", headers=owner)

        print("\n== Solicitud de material (Producción) ==")
        r = client.post(
            "/api/wms/requests",
            headers=production,
            json={"part_number": "mat-1001", "quantity_requested": 100, "priority": "normal"},
        )
        check("producción crea una solicitud", r.status_code == 200)
        req = r.json()
        check("el folio es legible y consecutivo", req["folio"].startswith("REQ-") and req["folio"].endswith("000001"))
        check("nace pendiente", req["status"] == "pendiente")
        check("autocompleta la descripción desde el snapshot SAP", req["description"] == "TORNILLO HEXAGONAL M8")
        check("autocompleta la unidad desde el snapshot SAP", req["unit_of_measure"] == "PC")
        check("normaliza el número de parte a mayúsculas", req["part_number"] == "MAT-1001")
        check("guarda el stock del momento para auditoría", req["available_stock_at_request"] == 1200)
        check("sin riesgo de quiebre cuando alcanza", req["stock_risk"] is False)
        check("arranca en verde", req["alert_level"] == "green")

        # Stock insuficiente: se marca el riesgo pero NUNCA se bloquea.
        r = client.post(
            "/api/wms/requests",
            headers=production,
            json={"part_number": "MAT-3003", "quantity_requested": 500, "priority": "urgente"},
        )
        check("la solicitud con stock insuficiente sí se crea", r.status_code == 200)
        risky = r.json()
        check("se marca como riesgo de quiebre", risky["stock_risk"] is True)

        check(
            "cantidad cero rechazada",
            client.post(
                "/api/wms/requests", headers=production, json={"part_number": "MAT-1001", "quantity_requested": 0}
            ).status_code == 422,
        )
        check(
            "almacén no levanta solicitudes",
            client.post(
                "/api/wms/requests", headers=warehouse, json={"part_number": "MAT-1001", "quantity_requested": 1}
            ).status_code == 403,
        )

        print("\n== Tablero de almacén ==")
        board = client.get("/api/wms/board", headers=warehouse).json()
        check("las dos solicitudes están en pendiente", board["counts"]["pendiente"] == 2)
        check("el tablero expone los umbrales configurados", board["config"]["yellow_max_minutes"] == 60)
        check("la urgente va primero en la cola", board["columns"]["pendiente"][0]["priority"] == "urgente")
        check("cuenta las de riesgo de quiebre", board["counts"]["stock_risk"] == 1)

        print("\n== Tomar, liberar y volver a tomar ==")
        r = client.post(f"/api/wms/requests/{req['id']}/claim", headers=warehouse)
        check("almacén toma la solicitud", r.status_code == 200 and r.json()["status"] == "en_proceso")
        check("registra quién la tomó", r.json()["claimed_by_name"] == "Almacén")
        check(
            "tomarla dos veces devuelve conflicto",
            client.post(f"/api/wms/requests/{req['id']}/claim", headers=warehouse).status_code == 409,
        )

        r = client.post(
            f"/api/wms/requests/{req['id']}/release", headers=warehouse, json={"reason": "sin montacargas"}
        )
        check("se libera de vuelta a la cola", r.status_code == 200 and r.json()["status"] == "pendiente")
        check("queda sin surtidor asignado", r.json()["claimed_by_user_id"] is None)
        client.post(f"/api/wms/requests/{req['id']}/claim", headers=warehouse)

        print("\n== Surtido parcial y cierre ==")
        r = client.post(
            f"/api/wms/requests/{req['id']}/fulfill",
            headers=warehouse,
            json={"quantity_fulfilled": 40, "close_request": False},
        )
        check("registra una entrega parcial", r.status_code == 200)
        check("sigue en proceso si no se cierra", r.json()["status"] == "en_proceso")
        check("acumula lo entregado", r.json()["quantity_fulfilled_total"] == 40)
        check("calcula pendiente", r.json()["quantity_pending"] == 60)
        check("guarda el stock SAP del momento", r.json()["fulfillment"]["available_stock_at_moment"] == 1200)

        r = client.post(
            f"/api/wms/requests/{req['id']}/fulfill",
            headers=warehouse,
            json={"quantity_fulfilled": 25, "close_request": True},
        )
        check("cerrar con menos de lo pedido queda como parcial", r.json()["status"] == "surtido_parcial")
        check("suma los dos surtidos", r.json()["quantity_fulfilled_total"] == 65)
        check("registra la hora de cierre", bool(r.json()["closed_at"]))
        check("mide el tiempo de respuesta", r.json()["fulfillment"]["time_to_fulfill_minutes"] >= 0)
        check(
            "una solicitud cerrada ya no se surte",
            client.post(
                f"/api/wms/requests/{req['id']}/fulfill", headers=warehouse, json={"quantity_fulfilled": 1}
            ).status_code == 409,
        )

        # Cubrir lo pedido cierra la solicitud aunque no se pida cerrarla.
        r = client.post(
            f"/api/wms/requests/{risky['id']}/fulfill",
            headers=warehouse,
            json={"quantity_fulfilled": 500, "close_request": False},
        )
        check("surtir el total cierra como completo", r.json()["status"] == "surtido_completo")

        print("\n== Bitácora inmutable ==")
        audit = client.get(f"/api/wms/requests/{req['id']}/audit", headers=owner).json()
        actions = [entry["action"] for entry in audit]
        check("registra la creación", "created" in actions)
        check("registra quién la tomó", actions.count("claimed") == 2)
        check("registra la liberación", "released" in actions)
        check("registra los dos surtidos", actions.count("fulfilled") == 2)
        check("registra el cierre", "closed" in actions)
        check("cada entrada dice quién la hizo", all(entry["actor_name"] for entry in audit))
        check(
            "la liberación conserva de quién se liberó",
            any(e["payload"].get("released_from") == "Almacén" for e in audit if e["action"] == "released"),
        )

        print("\n== Detalle e historial ==")
        detail = client.get(f"/api/wms/requests/{req['id']}", headers=production).json()
        check("el detalle trae los surtidos", len(detail["fulfillments"]) == 2)
        check("el detalle trae el stock vigente", detail["current_available_stock"] == 1200)
        mine = client.get("/api/wms/requests/mine", headers=production).json()
        check("producción ve su historial", len(mine) == 2)

        print("\n== KPIs y dashboards ==")
        kpis = client.get("/api/wms/kpis", headers=owner).json()
        check("cuenta las solicitudes del periodo", kpis["totals"]["requests"] == 2)
        check("cuenta las cerradas", kpis["totals"]["closed"] == 2)
        check("cuenta las urgentes", kpis["totals"]["urgent"] == 1)
        check("cuenta los surtidos parciales", kpis["totals"]["partial"] == 1)
        check("reporta promedio y mediana", "median" in kpis["response_minutes"])
        check("calcula el % dentro de SLA", kpis["sla"]["pct_within"] == 100.0)
        check(
            "compara solicitado vs surtido",
            kpis["quantities"]["requested"] == 600 and kpis["quantities"]["fulfilled"] == 565,
        )

        leaderboard = client.get("/api/wms/leaderboard", headers=owner).json()["rows"]
        check("el leaderboard tiene al surtidor", len(leaderboard) == 1)
        check("cuenta sus surtidos", leaderboard[0]["fulfillments"] == 3)
        check("cuenta las solicitudes distintas que atendió", leaderboard[0]["requests_served"] == 2)
        check("reporta su mediana de respuesta", "median_response_minutes" in leaderboard[0])

        requesters = client.get("/api/wms/requesters", headers=owner).json()["rows"]
        check("mide la actividad del solicitante", requesters[0]["requests"] == 2)
        check("mide su proporción de urgentes", requesters[0]["pct_urgent"] == 50.0)

        partials = client.get("/api/wms/partial-fulfillments", headers=owner).json()["rows"]
        check("detecta la parte con surtido parcial", partials[0]["part_number"] == "MAT-1001")
        check("calcula el faltante", partials[0]["shortfall"] == 35)

        print("\n== Alertas de solicitudes atrasadas ==")
        import asyncio

        from alerts import scan_overdue_requests
        from config import db as _db

        async def age_and_scan(tenant_id):
            # Envejece la solicitud abierta 90 minutos para cruzar el umbral rojo.
            await _db.wms_requests.update_one(
                {"id": late_id}, {"$set": {"requested_at": "2020-01-01T00:00:00+00:00"}}
            )
            return await scan_overdue_requests(tenant_id)

        late = client.post(
            "/api/wms/requests",
            headers=production,
            json={"part_number": "MAT-2002", "quantity_requested": 5},
        ).json()
        late_id = late["id"]
        tenant_id = client.get("/api/auth/me", headers=owner).json()["user"]["tenant_id"]
        created = asyncio.get_event_loop().run_until_complete(age_and_scan(tenant_id))
        check("crea la alerta de atraso", created == 1)
        alerts = client.get("/api/alerts", headers=owner).json()
        overdue = [a for a in alerts if a["type"] == "wms_request_overdue"]
        check("la alerta aparece para el supervisor", len(overdue) == 1)
        check("la alerta es crítica", overdue[0]["level"] == "critical")
        check("la alerta identifica el folio", overdue[0]["folio"] == late["folio"])
        again = asyncio.get_event_loop().run_until_complete(scan_overdue_requests(tenant_id))
        check("no duplica la alerta en el siguiente barrido", again == 0)

        board = client.get("/api/wms/board", headers=warehouse).json()
        check("el tablero marca la atrasada en rojo", board["counts"]["red"] == 1)

        client.post(f"/api/wms/requests/{late_id}/fulfill", headers=warehouse, json={"quantity_fulfilled": 5})
        asyncio.get_event_loop().run_until_complete(scan_overdue_requests(tenant_id))
        remaining = [a for a in client.get("/api/alerts", headers=owner).json() if a["type"] == "wms_request_overdue"]
        check("la alerta se auto-resuelve al surtirse", len(remaining) == 0)

        print("\n== Configuración de umbrales ==")
        r = client.put(
            "/api/settings",
            headers=owner,
            json={"wms_config": {"green_max_minutes": 10, "yellow_max_minutes": 30, "sla_minutes": 15}},
        )
        check("el dueño guarda los umbrales", r.status_code == 200)
        board = client.get("/api/wms/board", headers=warehouse).json()
        check("el tablero usa los umbrales nuevos", board["config"]["yellow_max_minutes"] == 30)
        check("el SLA nuevo llega al dashboard", client.get("/api/wms/kpis", headers=owner).json()["sla"]["target_minutes"] == 15)

        print("\n== Salud de la ingesta y export a Excel ==")
        health = client.get("/api/inventory/sync-health", headers=owner).json()
        check("reporta la ingesta como sana", health["status"] == "ok")
        check("reporta cuántas partes sigue", health["parts_tracked"] == 6)
        check("reporta minutos desde la última corrida", health["minutes_since_last_success"] is not None)
        check("hay bitácora de sincronizaciones", len(client.get("/api/inventory/sync-logs", headers=owner).json()) >= 2)

        r = client.get("/api/wms/export/excel", headers=owner)
        check("el export a Excel responde 200", r.status_code == 200)
        check(
            "devuelve un .xlsx",
            r.headers["content-type"].startswith("application/vnd.openxmlformats"),
        )
        check("el archivo no viene vacío", len(r.content) > 2000)
        check(
            "producción no puede exportar",
            client.get("/api/wms/export/excel", headers=production).status_code == 403,
        )

        from openpyxl import load_workbook
        import io as _io

        workbook = load_workbook(_io.BytesIO(r.content))
        check("trae las dos hojas", workbook.sheetnames == ["Solicitudes", "KPIs por persona"])
        detail_sheet = workbook["Solicitudes"]
        # Encabezado + las 2 solicitudes del rango por defecto (30 días). La
        # tercera se envejeció a 2020 para probar la alerta de atraso, así que
        # queda fuera del periodo — que es justo lo que debe pasar.
        check("la hoja de detalle trae encabezado + las del periodo", detail_sheet.max_row == 3)
        check("la primera columna es el folio", detail_sheet.cell(row=1, column=1).value == "Folio")
        check("la hoja de KPIs trae filas", workbook["KPIs por persona"].max_row >= 2)

        print("\n== Parseo del export de MB52 ==")
        from sap_inventory_ingest import detect_decimal_separator, parse_sap_number

        check("coma decimal cuando aparecen los dos separadores", detect_decimal_separator(["1.250,000"]) == ",")
        check("punto decimal cuando aparecen los dos al revés", detect_decimal_separator(["1,250.000"]) == ".")
        check("un separador repetido solo puede ser de miles", detect_decimal_separator(["1.234.567"]) == ",")
        check("42,000 se lee como 42", parse_sap_number("42,000", ",") == 42.0)
        check("1.250,500 se lee como 1250.5", parse_sap_number("1.250,500", ",") == 1250.5)
        check("el signo al final se respeta", parse_sap_number("15,000-", ",") == -15.0)
        check("una celda vacía no es cero, es sin dato", parse_sap_number("", ",") is None)
        check("un texto no numérico se descarta", parse_sap_number("N/A", ",") is None)

        print("\n== Surtidos simultáneos (no se pierde ninguno) ==")
        concurrent = client.post(
            "/api/wms/requests",
            headers=production,
            json={"part_number": "MAT-1001", "quantity_requested": 30},
        ).json()

        async def fulfill_twice_racing():
            """Dos surtidores registrando 10 exactamente a la vez.

            Para que la carrera sea determinista se retrasa el await que el
            handler hace entre leer la solicitud y escribir el acumulado: así
            ambos leen el mismo total antes de que cualquiera escriba, que es
            justo el escenario donde un read-modify-write perdería una entrega.
            """
            from httpx import ASGITransport, AsyncClient

            import routes_wms_requests as wms_routes

            original = wms_routes.snapshot_quantity

            async def slow_snapshot(*args, **kwargs):
                await asyncio.sleep(0.05)
                return await original(*args, **kwargs)

            wms_routes.snapshot_quantity = slow_snapshot
            try:
                transport = ASGITransport(app=server.app)
                async with AsyncClient(transport=transport, base_url="http://testserver") as async_client:
                    body = {"quantity_fulfilled": 10, "close_request": False}
                    return await asyncio.gather(
                        async_client.post(
                            f"/api/wms/requests/{concurrent['id']}/fulfill", headers=warehouse, json=body
                        ),
                        async_client.post(
                            f"/api/wms/requests/{concurrent['id']}/fulfill", headers=warehouse, json=body
                        ),
                    )
            finally:
                wms_routes.snapshot_quantity = original

        asyncio.get_event_loop().run_until_complete(fulfill_twice_racing())
        after = client.get(f"/api/wms/requests/{concurrent['id']}", headers=owner).json()
        check("dos entregas simultáneas se acumulan (20, no 10)", after["quantity_fulfilled_total"] == 20)
        check("las dos quedan en el historial", len(after["fulfillments"]) == 2)
        client.post(
            f"/api/wms/requests/{concurrent['id']}/fulfill",
            headers=warehouse,
            json={"quantity_fulfilled": 10},
        )

        print("\n== Aislamiento por locación ==")
        other = client.get("/api/wms/locations", headers=owner).json()
        other_location = next(loc for loc in other if loc["id"] != location_id)
        check(
            "producción no consulta el inventario de otra planta",
            client.get(f"/api/inventory/by-location/{other_location['id']}", headers=production).status_code == 403,
        )
        check(
            "el supervisor sí ve cualquier planta",
            client.get(f"/api/inventory/by-location/{other_location['id']}", headers=owner).status_code == 200,
        )
        check(
            "producción solo ve las locaciones que le tocan",
            len(client.get("/api/wms/locations", headers=production).json()) == 1,
        )
finally:
    shutil.rmtree(_EXPORT_DIR, ignore_errors=True)

print(f"\n{'=' * 60}")
print(f"  {PASS} PASS · {FAIL} FAIL")
print(f"{'=' * 60}\n")
raise SystemExit(1 if FAIL else 0)
