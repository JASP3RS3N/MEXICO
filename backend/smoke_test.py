"""End-to-end smoke test using an in-memory Mongo (mongomock-motor).

Patches the motor client before importing the app, then drives the real API
through FastAPI's TestClient: seeding, auth, RBAC, POS order, kitchen flow,
payment + inventory deduction, purchase orders and P&L.
Run: python3 smoke_test.py
"""
import mongomock_motor
import motor.motor_asyncio

# Patch motor with the in-memory implementation BEFORE the app imports config.
motor.motor_asyncio.AsyncIOMotorClient = mongomock_motor.AsyncMongoMockClient

from fastapi.testclient import TestClient  # noqa: E402
import server  # noqa: E402

from io import BytesIO  # noqa: E402

from openpyxl import Workbook, load_workbook  # noqa: E402

client = TestClient(server.app)
PASS, FAIL = 0, 0


def check(label, condition):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  PASS  {label}")
    else:
        FAIL += 1
        print(f"  FAIL  {label}")


def auth(username, password):
    r = client.post("/api/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, f"login {username} -> {r.status_code} {r.text}"
    return {"Authorization": f"Bearer {r.json()['token']}"}


def auth_pin(owner_headers, username):
    """Cashier/prep now log in by PIN (see commit 0c81b7e): find the user,
    issue them a PIN (seed users start with pin=None), then swap this
    device's session to that PIN via /auth/login-pin. Returns (headers, pin)
    so callers can also reuse the PIN itself (e.g. PinTagRequest bodies)."""
    users = client.get("/api/users", headers=owner_headers).json()
    target = next(u for u in users if u["username"] == username)
    r = client.post(f"/api/users/{target['id']}/regenerate-pin", headers=owner_headers)
    assert r.status_code == 200, f"regenerate-pin {username} -> {r.status_code} {r.text}"
    pin = r.json()["pin"]
    r = client.post("/api/auth/login-pin", headers=owner_headers, json={"pin": pin})
    assert r.status_code == 200, f"login-pin {username} -> {r.status_code} {r.text}"
    return {"Authorization": f"Bearer {r.json()['token']}"}, pin


with client:  # triggers startup (seed)
    print("\n== Auth & RBAC ==")
    owner = auth("dueno", "dueno123")
    cashier, cashier_pin = auth_pin(owner, "caja")
    prep, prep_pin = auth_pin(owner, "cocina")

    check("bad login rejected", client.post("/api/auth/login", json={"username": "dueno", "password": "x"}).status_code == 401)
    check("me returns owner role", client.get("/api/auth/me", headers=owner).json()["user"]["role"] == "owner")
    check("cashier cannot list users (403)", client.get("/api/users", headers=cashier).status_code == 403)
    check("cashier cannot see finances (403)", client.get("/api/finance/pnl", headers=cashier).status_code == 403)
    check("prep cannot see daily sales (403)", client.get("/api/finance/daily", headers=prep).status_code == 403)
    check("owner can see pnl", client.get("/api/finance/pnl", headers=owner).status_code == 200)

    print("\n== Seeded catalog ==")
    products = client.get("/api/products", headers=cashier).json()
    materials = client.get("/api/materials", headers=owner).json()
    check("products seeded", len(products) >= 8)
    check("materials seeded", len(materials) >= 10)
    check("products carry recipe cost", any(p.get("cost", 0) > 0 for p in products))

    print("\n== User management ==")
    r = client.post("/api/users", headers=owner, json={"username": "caja2", "name": "Cajera 2", "password": "pw123", "role": "cashier"})
    check("owner creates user", r.status_code == 200)
    check("duplicate user rejected", client.post("/api/users", headers=owner, json={"username": "caja2", "name": "x", "password": "y", "role": "cashier"}).status_code == 409)
    new_id = r.json()["id"]
    check("owner deletes user", client.delete(f"/api/users/{new_id}", headers=owner).status_code == 200)

    print("\n== Price edit (owner only) ==")
    p0 = products[0]
    check("cashier cannot change price (403)", client.patch(f"/api/products/{p0['id']}/price", headers=cashier, json={"price": 1}).status_code == 403)
    r = client.patch(f"/api/products/{p0['id']}/price", headers=owner, json={"price": 199.5})
    check("owner edits price", r.status_code == 200 and r.json()["price"] == 199.5)

    print("\n== POS order -> kitchen -> pay ==")
    brisket = next(p for p in products if "Brisket" in p["name"])
    refresco = next(p for p in products if p["name"] == "Refresco")
    mat_before = {m["id"]: m["current_stock"] for m in materials}

    r = client.post("/api/orders", headers=cashier, json={
        "customer_name": "Mesa 4",
        "items": [{"product_id": brisket["id"], "qty": 2}, {"product_id": refresco["id"], "qty": 1}],
    })
    check("cashier creates order", r.status_code == 200)
    order = r.json()
    check("order has sequential number", isinstance(order["order_number"], int))
    check("order total computed", order["total"] > 0)
    check("order starts pending", order["status"] == "pending")

    check("prep cannot create order (403)", client.post("/api/orders", headers=prep, json={"items": [{"product_id": brisket["id"], "qty": 1}]}).status_code == 403)

    # kitchen visibility
    kq = client.get("/api/kitchen", headers=prep).json()
    check("order visible in kitchen queue", any(o["order_number"] == order["order_number"] for o in kq))
    check("kitchen hides prices", all("price" not in i for o in kq for i in o["items"]))

    # public display (no auth)
    disp = client.get("/api/display", params={"tenant": "demo"}).json()
    check("public display shows order, no money", any(o["order_number"] == order["order_number"] for o in disp) and all("total" not in o for o in disp))

    # prep flow
    check("prep accepts order", client.post(f"/api/orders/{order['id']}/accept", headers=prep, json={"pin": prep_pin}).status_code == 200)
    check("cannot re-accept order", client.post(f"/api/orders/{order['id']}/accept", headers=prep, json={"pin": prep_pin}).status_code == 400)
    check("prep marks ready", client.post(f"/api/orders/{order['id']}/ready", headers=prep).status_code == 200)

    # payment + inventory deduction
    r = client.post(f"/api/orders/{order['id']}/pay", headers=cashier, json={"method": "efectivo", "amount_received": 1000})
    check("cashier pays order", r.status_code == 200)
    check("change computed", r.json()["change"] == round(1000 - order["total"], 2))
    check("cannot double-pay", client.post(f"/api/orders/{order['id']}/pay", headers=cashier, json={"method": "efectivo"}).status_code == 400)

    # propina (tip): el cambio se calcula sobre total + propina
    r = client.post("/api/orders", headers=cashier, json={"items": [{"product_id": refresco["id"], "qty": 1}]})
    check("cashier creates second order (tip test)", r.status_code == 200)
    tip_order = r.json()
    tip_amount = round(tip_order["total"] * 0.1, 2) or 5.0
    r = client.post(f"/api/orders/{tip_order['id']}/pay", headers=cashier, json={"method": "efectivo", "amount_received": 1000, "tip_amount": tip_amount})
    check("cashier pays order with tip", r.status_code == 200)
    check("change accounts for tip", r.json()["change"] == round(1000 - (tip_order["total"] + tip_amount), 2))
    got = client.get(f"/api/orders/{tip_order['id']}", headers=owner).json()
    check("order stores tip_amount", got.get("tip_amount") == tip_amount)

    # corrección de método de pago en orden ya cobrada (solo el método; con auditoría)
    r = client.post(f"/api/orders/{order['id']}/correct-payment-method", headers=cashier, json={"method": "tarjeta"})
    check("cashier corrects payment method on paid order", r.status_code == 200)
    corrected = r.json()
    check("payment method updated", corrected["payment_method"] == "tarjeta")
    check("audit keeps original method", corrected.get("original_payment_method") == "efectivo")
    check("audit records who/when", bool(corrected.get("payment_corrected_by_user_id")) and bool(corrected.get("payment_corrected_at")))
    check("correction does not change totals", corrected["total"] == order["total"] and corrected["amount_received"] == 1000.0)
    check("order stays paid (not reopened)", corrected.get("paid") is True)
    check("prep cannot correct payment method (403)", client.post(f"/api/orders/{order['id']}/correct-payment-method", headers=prep, json={"method": "efectivo"}).status_code == 403)

    # número de personas en mesa ya abierta (#32) — solo cashier/owner, con auditoría
    r = client.post("/api/orders", headers=cashier, json={
        "customer_name": "Mesa 7",
        "table": "7",
        "party_size": 2,
        "items": [{"product_id": refresco["id"], "qty": 1}],
    })
    check("cashier creates table order with party size", r.status_code == 200 and r.json().get("party_size") == 2)
    table_order = r.json()

    r = client.post(f"/api/orders/{table_order['id']}/party-size", headers=cashier, json={"party_size": 5})
    check("cashier changes party size on open table", r.status_code == 200 and r.json().get("party_size") == 5)
    check("party size audit records who/when", bool(r.json().get("party_size_changed_by_user_id")) and bool(r.json().get("party_size_changed_at")))

    check("owner can change party size", client.post(f"/api/orders/{table_order['id']}/party-size", headers=owner, json={"party_size": 4}).status_code == 200)
    check("prep cannot change party size (403)", client.post(f"/api/orders/{table_order['id']}/party-size", headers=prep, json={"party_size": 6}).status_code == 403)
    check("zero party size rejected (422)", client.post(f"/api/orders/{table_order['id']}/party-size", headers=cashier, json={"party_size": 0}).status_code == 422)

    r = client.post("/api/orders", headers=cashier, json={"items": [{"product_id": refresco["id"], "qty": 1}]})
    check("order without table rejects party-size (400)", client.post(f"/api/orders/{r.json()['id']}/party-size", headers=cashier, json={"party_size": 3}).status_code == 400)

    # orden ya cobrada → no se puede cambiar el número de personas
    check("paid order rejects party-size (400)", client.post(f"/api/orders/{order['id']}/party-size", headers=cashier, json={"party_size": 9}).status_code == 400)

    mats_after = {m["id"]: m["current_stock"] for m in client.get("/api/materials", headers=owner).json()}
    brisket_mat = next(m["material_id"] for m in brisket["recipe"] if True)
    check("inventory deducted after payment", mats_after[brisket_mat] < mat_before[brisket_mat])

    print("\n== Purchase orders ==")
    low_mat = materials[0]
    r = client.post("/api/purchase-orders", headers=owner, json={
        "supplier": "Proveedor X",
        "items": [{"material_id": low_mat["id"], "qty": 10, "unit_cost": 100}],
    })
    check("owner creates PO", r.status_code == 200)
    po = r.json()
    check("PO numbered", po["po_number"].startswith("OC-"))
    check("PO total", po["total"] == 1000.0)
    stock_before_po = next(m["current_stock"] for m in client.get("/api/materials", headers=owner).json() if m["id"] == low_mat["id"])
    # Control #5 — validaciones server-side al recibir (casos negativos). La PO
    # debe estar "ordered" para que el PUT recibido llegue a las nuevas reglas.
    neg_po = client.post("/api/purchase-orders", headers=owner, json={
        "supplier": "Proveedor X",
        "items": [{"material_id": low_mat["id"], "qty": 10, "unit_cost": 100}],
    }).json()
    check("neg PO marked ordered", client.put(f"/api/purchase-orders/{neg_po['id']}/status", headers=owner, json={"status": "ordered"}).status_code == 200)
    check("receive without physical_supplier rejected (422)", client.put(f"/api/purchase-orders/{neg_po['id']}/status", headers=owner, json={"status": "received"}).status_code == 422)

    neg_po2 = client.post("/api/purchase-orders", headers=owner, json={
        "supplier": "Proveedor X",
        "items": [{"material_id": low_mat["id"], "qty": 10, "unit_cost": 100}],
    }).json()
    check("neg PO2 marked ordered", client.put(f"/api/purchase-orders/{neg_po2['id']}/status", headers=owner, json={"status": "ordered"}).status_code == 200)
    check(">10% variance without reason rejected (422)", client.put(
        f"/api/purchase-orders/{neg_po2['id']}/status", headers=owner,
        json={
            "status": "received",
            "physical_supplier": "Proveedor X",
            "received_items": [{"material_id": low_mat["id"], "received_qty": 5}],
        },
    ).status_code == 422)

    check("PO marked ordered", client.put(f"/api/purchase-orders/{po['id']}/status", headers=owner, json={"status": "ordered"}).status_code == 200)
    check("receiving PO adds stock", client.put(f"/api/purchase-orders/{po['id']}/status", headers=owner, json={"status": "received", "physical_supplier": "Proveedor X"}).status_code == 200)
    stock_after_po = next(m["current_stock"] for m in client.get("/api/materials", headers=owner).json() if m["id"] == low_mat["id"])
    check("stock increased by received qty", stock_after_po == stock_before_po + 10)
    check("cannot re-receive PO", client.put(f"/api/purchase-orders/{po['id']}/status", headers=owner, json={"status": "received"}).status_code == 400)

    print("\n== Supplier quote template (#20) ==")
    r = client.post("/api/suppliers", headers=owner, json={"name": "Carnes del Norte"})
    check("owner creates supplier for template", r.status_code == 200)
    sup_id = r.json()["id"]

    mats_now = [m for m in client.get("/api/materials", headers=owner).json() if m.get("active", True)]
    r = client.get(f"/api/suppliers/{sup_id}/quote-template", headers=owner)
    check("template download ok (200)", r.status_code == 200)
    check("template is xlsx content type", "spreadsheetml" in r.headers.get("content-type", ""))

    wb = load_workbook(BytesIO(r.content))
    ws = wb["Cotización"]
    rows = list(ws.iter_rows())
    header_idx = next(i for i, row in enumerate(rows) if any(c.value == "Insumo" for c in row))
    data_rows = [row for row in rows[header_idx + 1:] if row[1].value]
    check("template lists every active material", len(data_rows) == len(mats_now))

    # Pre-fill: register an offering, regenerate, and verify the cost column picks it up.
    mat0 = mats_now[0]
    r = client.post(
        "/api/supplier-offerings",
        headers=owner,
        json={"supplier_id": sup_id, "material_id": mat0["id"], "cost_per_unit": 12.5, "min_order": 5, "lead_time_days": 3},
    )
    check("offering created for pre-fill test", r.status_code == 200)
    wb2 = load_workbook(BytesIO(client.get(f"/api/suppliers/{sup_id}/quote-template", headers=owner).content))
    ws2 = wb2["Cotización"]
    rows2 = list(ws2.iter_rows())
    h2 = next(i for i, row in enumerate(rows2) if any(c.value == "Insumo" for c in row))
    col_idx = {c.value: n for n, c in enumerate(rows2[h2])}
    prefilled = [row for row in rows2[h2 + 1:] if row[1].value == mat0["name"]]
    check("offering cost pre-filled in template", bool(prefilled) and float(prefilled[0][col_idx["Costo actual (MXN)"]].value) == 12.5)
    check("offering lead time pre-filled", bool(prefilled) and int(prefilled[0][col_idx["Lead time (días)"]].value) == 3)

    check("cashier cannot download template (403)", client.get(f"/api/suppliers/{sup_id}/quote-template", headers=cashier).status_code == 403)
    check("unknown supplier -> 404", client.get("/api/suppliers/nope/quote-template", headers=owner).status_code == 404)

    print("\n== Supplier quote import (#19) ==")
    XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    QUOTE_HEADERS = ["SKU", "Insumo", "Unidad", "Costo actual (MXN)", "Precio cotizado (MXN)", "Mínimo pedido", "Lead time (días)"]

    def build_quote(rows, header_row=6):
        wbq = Workbook()
        wsq = wbq.active
        wsq.title = "Cotización"
        for i in range(1, header_row):
            wsq.append(["SOLICITUD DE COTIZACIÓN"] if i == 1 else [])
        wsq.append(QUOTE_HEADERS)
        for r in rows:
            wsq.append(r)
        buf = BytesIO()
        wbq.save(buf)
        return buf.getvalue()

    def upload_quote(payload, name="cotizacion.xlsx", hdrs=owner):
        return client.post(f"/api/suppliers/{sup_id}/quote-import", headers=hdrs, files={"file": (name, payload, XLSX_MIME)})

    # 1) Round-trip: download the template (#20), fill prices and send it back.
    tpl_bytes = client.get(f"/api/suppliers/{sup_id}/quote-template", headers=owner).content
    wb3 = load_workbook(BytesIO(tpl_bytes))
    ws3 = wb3["Cotización"]
    h3 = next(i for i, row in enumerate(ws3.iter_rows(max_row=10), start=1) if any(c.value == "Insumo" for c in row))
    pcol = next(n + 1 for n, c in enumerate(ws3[h3]) if c.value and "Precio cotizado" in str(c.value))
    ws3.cell(row=h3 + 1, column=pcol, value=42.75)   # mat0: existing offering (12.5) -> updated
    ws3.cell(row=h3 + 1, column=6, value=8)
    ws3.cell(row=h3 + 1, column=7, value=5)
    ws3.cell(row=h3 + 2, column=pcol, value=18.90)   # second material: no offering -> created
    ws3.cell(row=h3 + 2, column=6, value=3)
    ws3.cell(row=h3 + 2, column=7, value=4)
    buf = BytesIO()
    wb3.save(buf)

    r = upload_quote(buf.getvalue(), name="cotizacion_devuelta.xlsx")
    check("quote import ok (200)", r.status_code == 200)
    res = r.json()
    imp_a = next((i for i in res.get("imported", []) if i["material_id"] == mat0["id"]), None)
    imp_b = next((i for i in res.get("imported", []) if i["material_id"] == mats_now[1]["id"]), None)
    check("existing offering updated with old/new cost", imp_a is not None and imp_a["action"] == "updated"
          and imp_a["old_cost_per_unit"] == 12.5 and imp_a["new_cost_per_unit"] == 42.75
          and imp_a["min_order"] == 8.0 and imp_a["lead_time_days"] == 5)
    check("missing offering created", imp_b is not None and imp_b["action"] == "created"
          and imp_b["new_cost_per_unit"] == 18.90 and imp_b["min_order"] == 3.0 and imp_b["lead_time_days"] == 4)
    check("only priced rows imported", len(res.get("imported", [])) == 2 and res.get("not_found") == [])
    check("blank-price rows reported as skipped", len(res.get("skipped", [])) == len(mats_now) - 2)

    # 2) Persistence: a fresh template must show the imported offering values.
    wb4 = load_workbook(BytesIO(client.get(f"/api/suppliers/{sup_id}/quote-template", headers=owner).content))
    ws4 = wb4["Cotización"]
    h4 = next(i for i, row in enumerate(ws4.iter_rows(max_row=10), start=1) if any(c.value == "Insumo" for c in row))
    rows_by_name = {str(rw[1].value): rw for rw in ws4.iter_rows(min_row=h4 + 1) if rw[1].value}
    ra, rb2 = rows_by_name[mat0["name"]], rows_by_name[mats_now[1]["name"]]
    check("updated offering persisted (cost/min/lead)", round(float(ra[3].value), 2) == 42.75 and float(ra[5].value) == 8 and int(ra[6].value) == 5)
    check("created offering persisted", round(float(rb2[3].value), 2) == 18.90 and float(rb2[5].value) == 3 and int(rb2[6].value) == 4)

    # 3) Hand-crafted file: unknown material + invalid price are reported, not applied.
    edge_mat = mats_now[2] if len(mats_now) > 2 else mat0
    r = upload_quote(build_quote([["", "Insumo que no existe", "", None, 99.0], [None, edge_mat["name"], "", None, "N/D"]]))
    check("unknown material reported in not_found", r.status_code == 200 and len(r.json().get("not_found", [])) == 1)
    check("invalid price reported in skipped", len(r.json().get("skipped", [])) == 1 and r.json().get("imported") == [])

    # 4) Header on row 1 (supplier removed the title rows) still parses.
    r = upload_quote(build_quote([[None, edge_mat["name"], "", None, 7.25]], header_row=1))
    check("header on first row accepted", r.status_code == 200 and len(r.json().get("imported", [])) == 1)

    # 5) Error cases.
    check("non-xlsx rejected (400)", upload_quote(b"hola mundo", name="cotizacion.txt").status_code == 400)
    check("corrupted xlsx rejected (400)", upload_quote(b"PK\x03\x04 garbage", name="roto.xlsx").status_code == 400)
    wb5 = Workbook()
    ws5 = wb5.active
    ws5.append(["Hola", "Mundo"])
    buf = BytesIO(); wb5.save(buf)
    check("file without quote headers rejected (400)", upload_quote(buf.getvalue()).status_code == 400)
    check("unknown supplier -> 404", client.post("/api/suppliers/nope/quote-import", headers=owner, files={"file": ("c.xlsx", tpl_bytes, XLSX_MIME)}).status_code == 404)
    check("cashier cannot import (403)", upload_quote(tpl_bytes, hdrs=cashier).status_code == 403)

    print("\n== Finance / P&L ==")
    pnl = client.get("/api/finance/pnl", headers=owner).json()
    check("pnl revenue > 0", pnl["revenue"] > 0)
    check("pnl has cogs", pnl["cogs"] >= 0)
    check("pnl gross_profit = revenue - cogs", pnl["gross_profit"] == round(pnl["revenue"] - pnl["cogs"], 2))
    check("pnl top_products populated", len(pnl["top_products"]) >= 1)

    client.post("/api/expenses", headers=owner, json={"category": "Renta", "description": "Renta local", "amount": 500})
    pnl2 = client.get("/api/finance/pnl", headers=owner).json()
    check("expenses reduce net profit", pnl2["net_profit"] == round(pnl2["gross_profit"] - pnl2["operating_expenses"], 2))

    daily = client.get("/api/finance/daily", headers=owner).json()
    check("daily has payment breakdown", len(daily["by_payment_method"]) >= 1)

    dash = client.get("/api/finance/dashboard", headers=owner).json()
    check("dashboard today block", "today" in dash and "month" in dash)

    print("\n== Cash movements / drawer open (#29) ==")
    tid = client.get("/api/auth/me", headers=owner).json()["user"]["tenant_id"]
    r = client.post("/api/cash-movements", headers=cashier, json={"type": "drawer_open", "reason": "Apertura de caja turno mañana"})
    check("cashier opens drawer (audited)", r.status_code == 200)
    mv = r.json()
    check("movement scoped to tenant", mv["tenant_id"] == tid)
    check("movement type is drawer_open", mv["type"] == "drawer_open")
    check("movement stores reason", mv["reason"] == "Apertura de caja turno mañana")
    check("movement records creator id", bool(mv.get("created_by_user_id")))
    check("movement records creator name", bool(mv.get("created_by_name")))
    check("movement has timestamp", bool(mv.get("created_at")))

    r = client.post("/api/cash-movements", headers=owner, json={"type": "drawer_open", "reason": "Apertura por dueño"})
    check("owner opens drawer (audited)", r.status_code == 200)

    check("prep cannot open drawer (403)", client.post("/api/cash-movements", headers=prep, json={"type": "drawer_open", "reason": "x"}).status_code == 403)
    check("missing reason rejected (400)", client.post("/api/cash-movements", headers=cashier, json={"type": "drawer_open"}).status_code == 400)
    check("blank reason rejected (400)", client.post("/api/cash-movements", headers=cashier, json={"type": "drawer_open", "reason": "   "}).status_code == 400)
    check("unsupported type rejected (400)", client.post("/api/cash-movements", headers=cashier, json={"type": "deposit", "reason": "x"}).status_code == 400)

    log = client.get("/api/cash-movements", headers=cashier).json()
    opens = [m for m in log if m["type"] == "drawer_open"]
    check("audit log lists both opens", len(opens) >= 2)

    print("\n== Cash auto-deposit on payment (#30) ==")
    moves = client.get("/api/cash-movements", headers=cashier).json()
    dep1 = next((m for m in moves if m["type"] == "deposit" and m.get("order_id") == order["id"]), None)
    check("cash payment created auto deposit", dep1 is not None)
    check("deposit amount equals total due", dep1["amount"] == round(order["total"], 2))
    check("deposit reason references the order", f"#{order['order_number']}" in dep1["reason"])
    dep_tip = next((m for m in moves if m["type"] == "deposit" and m.get("order_id") == tip_order["id"]), None)
    check("tip deposit includes tip amount", dep_tip is not None and dep_tip["amount"] == round(tip_order["total"] + tip_amount, 2))

    r = client.post("/api/orders", headers=cashier, json={"items": [{"product_id": refresco["id"], "qty": 1}]})
    card_order = r.json()
    check("card order paid", client.post(f"/api/orders/{card_order['id']}/pay", headers=cashier, json={"method": "tarjeta"}).status_code == 200)
    moves2 = client.get("/api/cash-movements", headers=cashier).json()
    check("card payment creates no cash movement", not any(m["type"] == "deposit" and m.get("order_id") == card_order["id"] for m in moves2))

print(f"\n==== RESULT: {PASS} passed, {FAIL} failed ====")
raise SystemExit(1 if FAIL else 0)
